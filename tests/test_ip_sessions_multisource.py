"""ip_sessions.xlsx: Logins aus ALLEN Quellen + Quellen-Abgleich (Betreuer-Wunsch).

Vorher wurden nur auth-SSH-Logins erfasst (fester event_type-Filter);
wtmp/wtmpdb-Logins und Journal-Sessions fielen durch -> In_Journal/In_Wtmp
blieben immer ✗. Jetzt fuehrt _detect_login alle vier Quellen zusammen.
"""
import sys
import tempfile
from pathlib import Path
from datetime import datetime, timezone

sys.path.insert(0, str(Path(__file__).parent.parent))

from models.pipeline_context import PipelineContext
from models.event import ForensicEvent
from utils.event_store import EventStore
from stages.stage14_export import _write_ip_sessions_excel, _detect_login

import pytest
pytest.importorskip('openpyxl')
from openpyxl import load_workbook

_T = lambda h, m=0: datetime(2024, 5, 1, h, m, tzinfo=timezone.utc)


def test_detect_login_alle_quellen():
    # auth SSH
    assert _detect_login(ForensicEvent(timestamp=_T(1), source='auth',
        event_type='ssh_login_success', message='Accepted password for a from 1.2.3.4',
        user='a', ip='1.2.3.4')) is not None
    # wtmp Konsole (ohne IP)
    u, ip, meth = _detect_login(ForensicEvent(timestamp=_T(1), source='wtmp',
        event_type='user_process', message='', user='netsec'))
    assert u == 'netsec' and ip == '' and 'Konsole' in meth
    # wtmpdb remote
    assert _detect_login(ForensicEvent(timestamp=_T(1), source='wtmpdb',
        event_type='user_process_login', message='', user='r', ip='9.9.9.9')) is not None
    # journald via Nachricht
    u, ip, meth = _detect_login(ForensicEvent(timestamp=_T(1), source='journald',
        event_type='system', message='Accepted publickey for bob from 5.6.7.8 port 22'))
    assert u == 'bob' and ip == '5.6.7.8'
    # boot ist KEIN Login
    assert _detect_login(ForensicEvent(timestamp=_T(1), source='wtmpdb',
        event_type='boot_time_login', message='')) is None


def _build():
    events = [
        ForensicEvent(timestamp=_T(14, 2), source='auth', event_type='ssh_login_success',
                      message='Accepted password for admin from 10.0.0.9', user='admin', ip='10.0.0.9'),
        ForensicEvent(timestamp=_T(14, 2), source='journald', event_type='system',
                      message='Accepted publickey for admin from 10.0.0.9 port 51000'),
        ForensicEvent(timestamp=_T(14, 2), source='wtmp', event_type='user_process',
                      message='user_process: User=admin Host=10.0.0.9', user='admin', ip='10.0.0.9'),
        ForensicEvent(timestamp=_T(15, 0), source='wtmpdb', event_type='user_process_login',
                      message='wtmpdb Login: User=root', user='root', ip='203.0.113.7'),
        ForensicEvent(timestamp=_T(16, 0), source='wtmp', event_type='user_process',
                      message='user_process: User=netsec Line=tty1', user='netsec'),
    ]
    case = Path(tempfile.mkdtemp()) / 'case'
    case.mkdir()
    db = case / 'events.db'
    with EventStore(db) as s:
        s.insert_events(events)
    ctx = PipelineContext(case_dir=case)
    ctx.events_db_path = db
    _write_ip_sessions_excel(ctx, case)
    wb = load_workbook(case / 'ip_sessions.xlsx')
    ws = wb['IP-Übersicht']
    rows = [r for r in ws.iter_rows(values_only=True)
            if r[0] and r[0] != 'IP' and 'IP-Session' not in str(r[0])]
    return wb, {r[0]: r for r in rows}


def test_quellen_abgleich_und_konsole():
    wb, d = _build()
    # 10.0.0.9 in auth + journal + wtmp belegt (Spalten 10/11/12 = idx 9/10/11)
    assert d['10.0.0.9'][9] == '✓' and d['10.0.0.9'][10] == '✓' and d['10.0.0.9'][11] == '✓'
    # 203.0.113.7 NUR in wtmpdb -> Diskrepanz (auth ✗, wtmpdb ✓)
    assert d['203.0.113.7'][9] == '✗' and d['203.0.113.7'][12] == '✓'
    # Konsolen-Login ohne IP sichtbar
    assert '(lokal/Konsole)' in d
    # Legende-Mappe vorhanden
    assert 'Legende' in wb.sheetnames
