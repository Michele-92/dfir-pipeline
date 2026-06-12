import csv
import json
import logging
import re
import time
from datetime import datetime, timezone
from pathlib import Path
from typing import Optional

from models.pipeline_context import PipelineContext
from stages.stage13_quality import evaluate_quality

try:
    from ki_text_generator import generate_all_critical_texts
    from report_builder import build_critical_report
    _KI_AVAILABLE = True
except ImportError:
    _KI_AVAILABLE = False

log = logging.getLogger(__name__)

# ── Farben (HEX → RGB tuple) ─────────────────────────────────────────────────
C_DARK_BLUE  = (0x1F/255, 0x4E/255, 0x79/255)
C_MID_BLUE   = (0x2E/255, 0x75/255, 0xB6/255)
C_LIGHT_BLUE = (0xD5/255, 0xE8/255, 0xF0/255)
C_DARK_GREY  = (0x40/255, 0x40/255, 0x40/255)
C_MID_GREY   = (0x59/255, 0x59/255, 0x59/255)
C_LIGHT_GREY = (0xF2/255, 0xF2/255, 0xF2/255)
C_WHITE      = (1, 1, 1)
C_GREEN      = (0x37/255, 0x56/255, 0x23/255)
C_GREEN_L    = (0xE2/255, 0xEF/255, 0xDA/255)
C_ORANGE     = (0x84/255, 0x3C/255, 0x0C/255)
C_ORANGE_L   = (0xFC/255, 0xE4/255, 0xD6/255)
C_RED        = (0xC0/255, 0x00/255, 0x00/255)
C_RED_L      = (0xFF/255, 0xCC/255, 0xCC/255)
C_YELLOW_L   = (0xFF/255, 0xEB/255, 0x9C/255)


def run(ctx: PipelineContext) -> PipelineContext:
    log.info('Stage 14: Export & Archivierung')
    case_dir = ctx.case_dir
    if not case_dir:
        log.error('Kein Case-Verzeichnis — Export übersprungen')
        return ctx

    _write_pipeline_report(ctx, case_dir)
    _write_autopsy_status(ctx, case_dir)
    _write_iocs_json(ctx, case_dir)
    _write_iocs_excel(ctx, case_dir)
    _write_antiforensics_json(ctx, case_dir)
    _write_activity_csv(ctx, case_dir)
    _export_mactime_package(ctx, case_dir)
    _export_forensic_findings_csv(ctx, case_dir)
    _export_forensic_findings_excel(ctx, case_dir)
    _write_ip_sessions_excel(ctx, case_dir)
    _write_reboot_sessions_excel(ctx, case_dir)
    _write_filtered_filesystem_timeline_excel(ctx, case_dir)
    _generate_report_pdf(ctx, case_dir)
    # NEU: moderner Bericht als ZUSAETZLICHE Datei (report.pdf bleibt bestehen)
    try:
        from stages.report_modern import build_modern_report
        build_modern_report(ctx, case_dir)
    except Exception as _e:
        log.warning(f'  Moderner Bericht uebersprungen: {_e}')
    _generate_coc_pdf(ctx, case_dir)
    _generate_critical_report(ctx, case_dir)
    _upload_timesketch(ctx, case_dir)

    log.info(f'  Export abgeschlossen → {case_dir}')
    if ctx.coc:
        ctx.coc.add_entry('stage_14', 'Export abgeschlossen')
    return ctx


def _generate_critical_report(ctx: PipelineContext, case_dir: Path) -> None:
    if not _KI_AVAILABLE:
        log.warning('  DFIR Critical Report übersprungen — ki_text_generator/report_builder nicht verfügbar')
        return
    critical = [f for f in ctx.forensic_findings if f.severity == 'CRITICAL']
    if not critical:
        log.info('  DFIR Critical Report übersprungen — keine CRITICAL-Befunde')
        return
    try:
        log.info(f'  Generiere KI-Texte für {len(critical)} CRITICAL-Befunde...')
        ki_texte_map = generate_all_critical_texts(
            findings       = ctx.forensic_findings,
            delay_seconds  = 0.3,
        )
        out = case_dir / 'DFIR_Critical_Report.pdf'
        build_critical_report(
            findings     = ctx.forensic_findings,
            ki_texte_map = ki_texte_map,
            ctx          = ctx,
            output_path  = out,
        )
        log.info(f'  DFIR_Critical_Report.pdf → {out}')
    except Exception as e:
        log.warning(f'  DFIR Critical Report Fehler: {e}')


# ── pipeline_report.json ─────────────────────────────────────────────────────

def _safe_duration(start_time) -> int:
    """Minuten seit start_time — robust gegen naive/aware-Mischung."""
    if start_time is None:
        return 0
    try:
        now = datetime.now(timezone.utc) if start_time.tzinfo else datetime.now()
        return (now - start_time).seconds // 60
    except Exception:
        return 0


def _to_utc(dt) -> datetime:
    """Normalisiert datetime fuer Sortierschluessel — immer timezone-aware UTC."""
    if dt is None:
        return datetime.min.replace(tzinfo=timezone.utc)
    if dt.tzinfo is None:
        return dt.replace(tzinfo=timezone.utc)
    return dt


def _write_pipeline_report(ctx: PipelineContext, case_dir: Path) -> None:
    duration = _safe_duration(ctx.start_time)
    report   = {
        'meta': {
            'case_id':          case_dir.name,
            'created':          datetime.utcnow().isoformat() + 'Z',
            'pipeline_version': '3.0',
            'duration_minutes': duration,
        },
        'input': {
            'disk_image': str(ctx.disk_image_path),
            'ram_dump':   str(ctx.ram_dump_path) if ctx.ram_dump_path else None,
            'sha256':     ctx.sha256,
            'md5':        ctx.md5,
            'size_gb':    round(ctx.file_size_gb, 2),
            'file_type':  ctx.file_type,
        },
        'system_profile': {
            'os_family':  ctx.os_family,
            'os_name':    ctx.os_name,
            'kernel':     ctx.kernel_version,
            'hostname':   ctx.hostname,
            'timezone':   ctx.timezone,
        },
        'statistics': {
            'total_log_lines':    ctx.total_log_lines,
            'parsed_events':      ctx.parsed_events,
            'anomalies_found':    len(ctx.anomalies),
            'iocs_found':         len(ctx.iocs),
            'mitre_techniques':   len(ctx.mitre_hits),
            'antiforensics_hits': len(ctx.antiforensics_hits),
        },
        'iocs': [
            {'type': ioc.type, 'value': ioc.value, 'source': ioc.source}
            for ioc in ctx.iocs[:100]
        ],
        'mitre_hits': ctx.mitre_hits,
        'antiforensics': ctx.antiforensics_hits,
        'stage_status': ctx.stage_status,
        'quality': evaluate_quality(ctx),
        'ioc_quality': ctx.ioc_quality,
    }
    out = case_dir / 'pipeline_report.json'
    out.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding='utf-8')
    log.info(f'  pipeline_report.json → {out}')


def _write_autopsy_status(ctx: PipelineContext, case_dir: Path) -> None:
    status = {
        'ran':    ctx.autopsy_ran,
        'reason': ctx.autopsy_reason,
        'results_count': len(ctx.autopsy_results.get('files', [])),
    }
    out = case_dir / 'autopsy_status.json'
    out.write_text(json.dumps(status, indent=2), encoding='utf-8')


def _write_iocs_json(ctx: PipelineContext, case_dir: Path) -> None:
    data = [
        {
            'type':       ioc.type,
            'value':      ioc.value,
            'source':     ioc.source,
            'timestamp':  ioc.timestamp.isoformat() if ioc.timestamp else None,
            'context':    ioc.context,
        }
        for ioc in ctx.iocs
    ]
    out = case_dir / 'iocs.json'
    out.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding='utf-8')
    log.info(f'  iocs.json → {out}')


def _write_antiforensics_json(ctx: PipelineContext, case_dir: Path) -> None:
    out = case_dir / 'antiforensics.json'
    out.write_text(
        json.dumps(ctx.antiforensics_hits, indent=2, ensure_ascii=False, default=str),
        encoding='utf-8'
    )
    log.info(f'  antiforensics.json → {out}')


# ── Activity CSV (Logins, Reboots, Crashes) ──────────────────────────────────

# Schicht 1: Strukturierte Event-Types aus Parsern (primär, zuverlässig)
_REBOOT_EVENT_TYPES = {'boot', 'system_boot', 'reboot', 'shutdown', 'system_shutdown',
                       'service_start', 'kernel_start'}
_LOGIN_EVENT_TYPES  = {'ssh_login_success', 'ssh_login_fail', 'ssh_invalid_user',
                       'sudo_command', 'user_login', 'console_login', 'su_session',
                       'pam_session', 'auth_success', 'auth_failure'}
_CRASH_EVENT_TYPES  = {'kernel_panic', 'oom_kill', 'segfault', 'service_crash',
                       'kernel_error', 'oom_killer'}

# Schicht 2: Keyword-Fallback für generische Events (nur HIGH/CRITICAL)
_REBOOT_KEYWORDS = ['system boot', 'kernel command line', 'reached target basic system',
                    'systemd: starting', 'reboot', 'shutdown']
_LOGIN_KEYWORDS  = ['session opened', 'accepted password', 'accepted publickey',
                    'new session', 'logged in']
_CRASH_KEYWORDS  = ['kernel panic', 'oom killer', 'segfault', 'out of memory',
                    'killed process', 'call trace']

# ── Reboot-Session-Konstanten ─────────────────────────────────────────────────
_SHUTDOWN_EVENT_TYPES = {'shutdown', 'system_shutdown'}
_BOOT_EVENT_TYPES     = {'boot', 'system_boot', 'kernel_start'}
_SHUTDOWN_KEYWORDS    = ['shutting down', 'halt', 'poweroff', 'initiating shutdown',
                         'reached target shutdown']
_BOOT_KEYWORDS_REBOOT = ['system boot', 'kernel command line', 'linux version',
                         'reached target basic system']

# ── IP-Session-Konstanten ─────────────────────────────────────────────────────
_SUCCESS_LOGIN_TYPES = {
    'ssh_login_success', 'auth_success', 'pam_session',
    'user_login', 'console_login', 'su_session',
}

# Login-Indikatoren im Nachrichtentext (journald/syslog: event_type ist dort
# generisch 'system', der Login steckt nur in der Message).
_LOGIN_MSG_RE = [
    re.compile(r'Accepted (?:password|publickey|keyboard-interactive\S*) '
               r'for (?P<user>\S+) from (?P<ip>[\d.]+|[0-9a-fA-F:]+)'),
    re.compile(r'New session \d+ of user (?P<user>\S+)'),
    re.compile(r'session opened for user (?P<user>\S+)'),
    re.compile(r'login: (?P<user>\S+)'),
]


def _detect_login(event):
    """Erkennt erfolgreiche Logins ueber ALLE Quellen (auth, journald, wtmp,
    wtmpdb). Gibt (user, ip, methode) zurueck oder None.

    Behebt: vorher wurden nur auth-SSH-Logins erfasst (fester event_type-Filter)
    — wtmp/wtmpdb-Logins und Journal-Sessions fielen durch, der Quellen-Abgleich
    lief leer. Jetzt:
      - auth/ssh:  explizite Login-Event-Typen (SSH ueber Passwort/Key)
      - wtmp/utmp: user_process = interaktive Session (Remote ODER Konsole)
      - wtmpdb:    *_login (Y2038-sicherer wtmp-Nachfolger)
      - journald/syslog: Login per Nachrichten-Muster (inkl. Konsole/CLI)
    """
    et   = (event.event_type or '').lower()
    src  = (event.source or '').lower()
    user = (event.user or '').strip()
    ip   = (event.ip or '').strip()

    if et in _SUCCESS_LOGIN_TYPES:
        return (user, ip, 'SSH-Schluessel' if 'publickey' in (event.message or '')
                else 'SSH-Passwort' if 'ssh' in et else 'PAM/Auth')
    if src in ('wtmp', 'utmp') and et == 'user_process':
        return (user, ip, 'Remote (SSH/Netzwerk)' if ip else 'Konsole/lokal')
    if src == 'wtmpdb' and et.endswith('_login') and not any(
            x in et for x in ('boot', 'run_level', 'empty')):
        return (user, ip, 'Remote (SSH/Netzwerk)' if ip else 'Konsole/lokal')
    if src in ('journald', 'syslog') or 'messages' in src:
        msg = event.message or ''
        for pat in _LOGIN_MSG_RE:
            m = pat.search(msg)
            if m:
                gd = m.groupdict()
                u  = user or gd.get('user', '') or ''
                i  = ip or gd.get('ip', '') or ''
                meth = ('SSH (Journal)' if 'Accepted' in msg
                        else 'Session (Journal)')
                return (u, i, meth)
    return None

# ── Filesystem-Timeline-Konstanten ────────────────────────────────────────────
_INTERESTING_PREFIXES = (
    '/home/', '/root/', '/tmp/', '/var/tmp/', '/dev/shm/',
    '/etc/', '/usr/local/', '/opt/', '/srv/', '/var/www/',
    '/run/', '/var/spool/',
)
_NOISE_PREFIXES = (
    '/usr/lib/', '/usr/share/', '/lib/', '/lib64/',
    '/proc/', '/sys/', '/dev/', '/run/lock/', '/snap/',
)
_PATH_CATEGORY_MAP = [
    ('/tmp/',       'Staging',   'FFF5F5', 'FFEDED'),
    ('/dev/shm/',   'Staging',   'FFF5F5', 'FFEDED'),
    ('/var/tmp/',   'Staging',   'FFF5F5', 'FFEDED'),
    ('/home/',      'User-Dir',  'FFF4DC', 'FFEAC8'),
    ('/root/',      'User-Dir',  'FFF4DC', 'FFEAC8'),
    ('/run/',       'Runtime',   'FFF4DC', 'FFEAC8'),
    ('/etc/',       'Config',    'FFFDE7', 'FFF9C4'),
    ('/usr/local/', 'Custom',    'F0F7FF', 'E3F0FF'),
    ('/opt/',       'Custom',    'F0F7FF', 'E3F0FF'),
    ('/srv/',       'Web/Data',  'F0FFF4', 'DFFFEB'),
    ('/var/www/',   'Web-Root',  'F0FFF4', 'DFFFEB'),
    ('/var/spool/', 'Spool',     'F8F0FF', 'F0E3FF'),
]


def _evidence_source(event) -> str:
    """Quelle eines Events mit Image-Praefix im Fall-Modus.
    '[webserver.E01] /var/log/auth.log' bzw. nur die Quelle bei 1 Image."""
    ev  = getattr(event, 'evidence', '') or ''
    src = getattr(event, 'orig_path', '') or getattr(event, 'source', '') or ''
    return f'[{ev}] {src}' if ev else src


def _classify_event(event) -> Optional[str]:
    """Klassifiziert Event als REBOOT/LOGIN/CRASH — strukturiert zuerst, Keyword-Fallback."""
    et = (event.event_type or '').lower()
    if et in _REBOOT_EVENT_TYPES: return 'REBOOT'
    if et in _LOGIN_EVENT_TYPES:  return 'LOGIN'
    if et in _CRASH_EVENT_TYPES:  return 'CRASH'
    # Keyword-Fallback nur für HIGH/CRITICAL Events
    if getattr(event, 'severity', '') in ('high', 'critical', 'error'):
        msg = event.message.lower()
        if any(kw in msg for kw in _CRASH_KEYWORDS):  return 'CRASH'
        if any(kw in msg for kw in _REBOOT_KEYWORDS): return 'REBOOT'
    # Keyword-Fallback für alle Severity bei Login/Reboot
    msg = event.message.lower()
    if any(kw in msg for kw in _LOGIN_KEYWORDS):  return 'LOGIN'
    if any(kw in msg for kw in _REBOOT_KEYWORDS): return 'REBOOT'
    return None


def _classify_reboot_event(event) -> Optional[str]:
    """Unterscheidet Boot- von Shutdown-Events innerhalb der REBOOT-Klasse."""
    et = (event.event_type or '').lower()
    if et in _BOOT_EVENT_TYPES:
        return 'boot'
    if et in _SHUTDOWN_EVENT_TYPES:
        return 'shutdown'
    msg = event.message.lower()
    if any(kw in msg for kw in _BOOT_KEYWORDS_REBOOT):
        return 'boot'
    if any(kw in msg for kw in _SHUTDOWN_KEYWORDS):
        return 'shutdown'
    return None


def _ip_type(ip: str) -> str:
    """Klassifiziert IP als Extern / Intern / Loopback."""
    if ip in ('127.0.0.1', '::1', 'localhost', '0.0.0.0'):
        return 'Loopback'
    if (ip.startswith('10.') or ip.startswith('192.168.')
            or ip.startswith('127.')
            or any(ip.startswith(f'172.{i}.') for i in range(16, 32))):
        return 'Intern'
    return 'Extern'


def _fmt_duration(td) -> str:
    """timedelta → 'Xd Xh Xm' oder '< 1 Min'."""
    total_sec = int(abs(td.total_seconds()))
    if total_sec < 60:
        return '< 1 Min'
    days    = total_sec // 86400
    hours   = (total_sec % 86400) // 3600
    minutes = (total_sec % 3600)  // 60
    parts = []
    if days:    parts.append(f'{days}d')
    if hours:   parts.append(f'{hours}h')
    if minutes: parts.append(f'{minutes}m')
    return ' '.join(parts) or '< 1 Min'


def _path_category(fp: str) -> tuple:
    """Gibt (Kategorie, fill_hex_A, fill_hex_B) für einen Dateipfad zurück."""
    for prefix, cat, fa, fb in _PATH_CATEGORY_MAP:
        if fp.startswith(prefix):
            return cat, fa, fb
    return 'Sonstige', 'FFFFFF', 'F2F5F9'


def _write_activity_csv(ctx: PipelineContext, case_dir: Path) -> None:
    """Exportiert Logins, Reboots und Crashes — kombiniert + 3 separate CSVs."""
    if not ctx.events_db_path or not ctx.events_db_path.exists():
        log.warning('  activity CSVs: keine events.db vorhanden')
        return
    from utils.event_store import EventStore
    relevant = []
    with EventStore(ctx.events_db_path) as store:
        for event in store.iter_events():
            classified = _classify_event(event)
            if classified:
                relevant.append((event, classified))

    relevant.sort(key=lambda x: x[0].timestamp)

    # Kombinierte Timeline
    out = case_dir / 'activity_timeline.csv'
    with open(out, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(['Timestamp_UTC', 'Event_Type', 'User', 'IP', 'Source', 'Details'])
        for event, et in relevant:
            writer.writerow([event.timestamp.isoformat(), et,
                             event.user or '', event.ip or '',
                             event.source, event.message[:300]])

    # 3 separate CSVs
    _write_event_csvs(relevant, case_dir)
    log.info(f'  activity_timeline.csv + 3 separate CSVs → {case_dir}  ({len(relevant)} Einträge)')


def _write_event_csvs(relevant: list, case_dir: Path) -> None:
    """Schreibt system_reboots.csv, login_events.csv, system_crashes.csv."""
    configs = {
        'REBOOT': ('system_reboots.csv',
                   ['Timestamp_UTC', 'Source', 'Details']),
        'LOGIN':  ('login_events.csv',
                   ['Timestamp_UTC', 'User', 'IP', 'Method', 'Source', 'Details']),
        'CRASH':  ('system_crashes.csv',
                   ['Timestamp_UTC', 'Severity', 'Source', 'Details']),
    }
    handles = {}
    writers = {}
    for et, (fname, header) in configs.items():
        f = open(case_dir / fname, 'w', newline='', encoding='utf-8')
        handles[et] = f
        writers[et] = csv.writer(f)
        writers[et].writerow(header)

    for event, et in relevant:
        if et == 'REBOOT':
            writers['REBOOT'].writerow([event.timestamp.isoformat(),
                                        event.source, event.message[:200]])
        elif et == 'LOGIN':
            method = event.event_type or ''
            writers['LOGIN'].writerow([event.timestamp.isoformat(),
                                       event.user or '', event.ip or '',
                                       method, event.source, event.message[:200]])
        elif et == 'CRASH':
            sev = getattr(event, 'severity', '')
            writers['CRASH'].writerow([event.timestamp.isoformat(),
                                       sev, event.source, event.message[:200]])

    for f in handles.values():
        f.close()


# ── MACtime CSV-Paket ────────────────────────────────────────────────────────

def _export_mactime_package(ctx: PipelineContext, case_dir: Path) -> None:
    """Exportiert MACtime-Events aus DuckDB als standalone CSV-Paket."""
    if not ctx.events_db_path or not ctx.events_db_path.exists():
        return
    from utils.event_store import EventStore
    out = case_dir / 'filesystem_timeline.csv'
    count = 0
    with EventStore(ctx.events_db_path) as store:
        with open(out, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerow(['Timestamp_UTC', 'MACB_Type', 'File_Path', 'Severity', 'Message'])
            for event in store.iter_events():
                if event.source != 'mactime':
                    continue
                writer.writerow([
                    event.timestamp.isoformat(),
                    event.event_type,
                    event.file_path or '',
                    event.severity,
                    event.message[:300],
                ])
                count += 1
    log.info(f'  filesystem_timeline.csv → {out}  ({count:,} MACtime-Events)')


# ── Forensic Findings CSV ────────────────────────────────────────────────────

def _export_forensic_findings_csv(ctx: PipelineContext, case_dir: Path) -> None:
    """Exportiert Stage-8.5-Befunde als forensic_findings.csv."""
    findings = getattr(ctx, 'forensic_findings', None)
    if not findings:
        return
    out = case_dir / 'forensic_findings.csv'
    with open(out, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow([
            'Timestamp_UTC', 'Severity', 'Check_Type',
            'Datei', 'Befund', 'Kontext_Stage6',
        ])
        for finding in findings:
            ts_str = (
                finding.anomaly_time.strftime('%Y-%m-%dT%H:%M:%SZ')
                if finding.anomaly_time else ''
            )
            evidence_str = ' | '.join(
                f"{e.get('time','')} [{e.get('source','')}] {e.get('message','')[:80]}"
                for e in (finding.evidence or [])[:3]
            )
            writer.writerow([
                ts_str,
                finding.severity,
                finding.rule,
                finding.file,
                finding.description,
                evidence_str,
            ])
    log.info(f'  forensic_findings.csv → {out}  ({len(findings)} Befunde)')


# ── Forensic Findings Excel ──────────────────────────────────────────────────

def _export_forensic_findings_excel(ctx: PipelineContext, case_dir: Path) -> None:
    """Exportiert Stage-8.5-Befunde als gestyltes Excel (Style-Guide v2)."""
    findings = getattr(ctx, 'forensic_findings', None)
    if not findings:
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning('openpyxl nicht installiert — Excel-Export übersprungen')
        return

    # ── Style-Helfer ─────────────────────────────────────────────
    def _fill(hex6):
        return PatternFill(fill_type='solid', fgColor=hex6.lstrip('#'))

    def _font(bold=False, color='000000', size=9):
        return Font(name='Arial', bold=bold, color=color.lstrip('#'), size=size)

    def _align(h='left', v='top', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

    _side   = Side(style='thin', color='C8D5E8')
    _BORDER = Border(left=_side, right=_side, top=_side, bottom=_side)

    # Fills
    F_BANNER = _fill('0D1B2A')
    F_HEADER = _fill('1B3A5C')
    F_W      = _fill('FFFFFF')
    F_ALT    = _fill('F2F5F9')

    # Zeilen-Hintergrund alternierend nach Severity
    ROW_BG = {
        'CRITICAL': [_fill('FFF5F5'), _fill('FFEDED')],
        'HIGH':     [_fill('FFFBF0'), _fill('FFF4DC')],
        'MEDIUM':   [_fill('FAFCFF'), _fill('F2F7FF')],
        'LOW':      [F_W, F_ALT],
    }
    # Severity-Badge (zentriert, fett, weiß)
    BADGE = {
        'CRITICAL': _fill('A32D2D'),
        'HIGH':     _fill('BA7517'),
        'MEDIUM':   _fill('5F5E5A'),
        'LOW':      _fill('3A6B3A'),
    }

    # Sortierung: CRITICAL → HIGH → MEDIUM → LOW, dann Timestamp aufsteigend
    # _to_utc() normalisiert naive/aware-Mischung → kein TypeError beim Vergleich
    _sev_ord = {'CRITICAL': 0, 'HIGH': 1, 'MEDIUM': 2, 'LOW': 3}
    sorted_f = sorted(findings, key=lambda f: (
        _sev_ord.get(f.severity, 4),
        _to_utc(f.anomaly_time),
    ))

    wb  = Workbook()
    HDR = ['Timestamp_UTC', 'Severity', 'Check_Type', 'Datei', 'Befund', 'Kontext_Stage6']
    WID = [22, 12, 28, 45, 65, 60]
    NC  = len(HDR)
    LC  = get_column_letter(NC)

    # ── Sheet 1: Findings ────────────────────────────────────────
    ws          = wb.active
    ws.title    = 'Findings'
    ws.sheet_view.showGridLines = False

    # Zeile 1: Titel-Banner
    ws.merge_cells(f'A1:{LC}1')
    b = ws.cell(1, 1, 'DFIR Forensic Findings — Stage 8.5 Timeline-Analyse')
    b.font      = _font(bold=True, color='FFFFFF', size=12)
    b.fill      = F_BANNER
    b.alignment = _align(h='center', v='center')
    ws.row_dimensions[1].height = 24

    # Zeile 2: Spalten-Header
    for col, (h, w) in enumerate(zip(HDR, WID), 1):
        c = ws.cell(2, col, h)
        c.font      = _font(bold=True, color='FFFFFF', size=10)
        c.fill      = F_HEADER
        c.alignment = _align(h='center', v='center')
        c.border    = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20

    ws.freeze_panes        = 'A3'
    ws.auto_filter.ref     = f'A2:{LC}{len(sorted_f) + 2}'

    # Datenzeilen
    _sev_cnt = {'CRITICAL': 0, 'HIGH': 0, 'MEDIUM': 0, 'LOW': 0}
    for ri, f in enumerate(sorted_f, 3):
        ts_str = (
            f.anomaly_time.strftime('%Y-%m-%dT%H:%M:%SZ')
            if f.anomaly_time else ''
        )
        ev_str = ' | '.join(
            f"{e.get('time','')} [{e.get('source','')}] {e.get('message','')[:80]}"
            for e in (f.evidence or [])[:3]
        )
        vals = [
            ts_str, f.severity, f.rule,
            f.file, f.description, ev_str,
        ]
        sev  = f.severity
        cnt  = _sev_cnt.get(sev, 0)
        rbg  = ROW_BG.get(sev, [F_W, F_ALT])[cnt % 2]
        _sev_cnt[sev] = cnt + 1

        for col, val in enumerate(vals, 1):
            c        = ws.cell(ri, col, val)
            c.border = _BORDER
            if col == 2:                          # Severity-Badge
                c.fill      = BADGE.get(sev, F_HEADER)
                c.font      = _font(bold=True, color='FFFFFF', size=9)
                c.alignment = _align(h='center', v='center')
            else:
                c.fill      = rbg
                c.font      = _font(size=9)
                c.alignment = _align(h='left', v='top', wrap=(col >= 5))

        ws.row_dimensions[ri].height = 28 if sev == 'CRITICAL' else 18

    # ── Sheet 2: Summary ─────────────────────────────────────────
    ws2 = wb.create_sheet('Summary')
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions['A'].width = 38
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 55
    ws2.column_dimensions['D'].width = 22

    # Banner
    ws2.merge_cells('A1:D1')
    b2 = ws2.cell(1, 1, 'DFIR Forensic Findings — Zusammenfassung')
    b2.font      = _font(bold=True, color='FFFFFF', size=12)
    b2.fill      = F_BANNER
    b2.alignment = _align(h='center', v='center')
    ws2.row_dimensions[1].height = 24

    ts_c = ws2.cell(2, 1, f'Generiert: {datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")}')
    ts_c.font = _font(size=9, color='5F5E5A')

    # Severity-Tabelle (mit COUNTIF)
    r = 4
    for col, lbl in enumerate(['Severity', 'Anzahl'], 1):
        c = ws2.cell(r, col, lbl)
        c.font      = _font(bold=True, color='FFFFFF', size=10)
        c.fill      = F_HEADER
        c.alignment = _align(h='center', v='center')
        c.border    = _BORDER
    ws2.row_dimensions[r].height = 20

    for i, sev in enumerate(['CRITICAL', 'HIGH', 'MEDIUM', 'LOW'], r + 1):
        c1 = ws2.cell(i, 1, sev)
        c2 = ws2.cell(i, 2, f'=COUNTIF(Findings!B:B,"{sev}")')
        c1.fill      = BADGE.get(sev, F_HEADER)
        c1.font      = _font(bold=True, color='FFFFFF', size=9)
        c1.alignment = _align(h='center', v='center')
        c2.fill      = ROW_BG.get(sev, [F_W])[0]
        c2.font      = _font(bold=True, size=9)
        c2.alignment = _align(h='center', v='center')
        for c in [c1, c2]:
            c.border = _BORDER
        ws2.row_dimensions[i].height = 18

    # Check-Typ-Tabelle mit Legende (mit COUNTIF)
    r2 = r + 7
    for col, lbl in enumerate(['Check-Typ', 'Anzahl', 'Bedeutung', 'Bewertung'], 1):
        c = ws2.cell(r2, col, lbl)
        c.font      = _font(bold=True, color='FFFFFF', size=10)
        c.fill      = F_HEADER
        c.alignment = _align(h='center', v='center')
        c.border    = _BORDER
    ws2.row_dimensions[r2].height = 20

    LEGEND = [
        ('night_activity',             'Systemaktivität 00:00–05:00 Uhr',                        '⚠ Verdächtig / oft Cronjob'),
        ('timestomping_C_gt_M',        'ctime > mtime, da ctime nicht fälschbar',                '🔴 Timestomping-Spur'),
        ('timestomping_M_lt_B',        'mtime < birthtime, oft dpkg FP, in /home kritisch',      '🔴 Timestomping / FP möglich'),
        ('timestamp_before_2000',      'Timestamp vor 2000 auf modernem System unrealistisch',    '⚠ Verdächtig'),
        ('staging_area',               'Datei in /tmp /dev/shm /var/tmp — Angreifer-Ablagepfad', '🔴 Kritischer Pfad'),
        ('activity_burst_system_path', 'Viele Systemdateien in kurzer Zeit modifiziert',          '🚨 Kritisch'),
        ('deleted_system_file',        'Gelöschte Systemdatei — Inode noch vorhanden',            '🔴 Kritisch'),
        ('cve_time_window',            'CVE ±30 Min neben Systemdatei-Modifikation gefunden',     '🚨 Kritisch'),
        ('sorter_extension_mismatch',  'Dateiendung passt nicht zu Magic-Bytes — Tarnung',        '🔴 Kritisch'),
        ('sorter_exec_in_staging',     'Executable in Staging-Bereich erkannt',                   '🚨 Kritisch'),
    ]

    for i, (check, desc, rating) in enumerate(LEGEND, r2 + 1):
        rbg = F_W if i % 2 == 0 else F_ALT
        for col, val in enumerate([
            check,
            f'=COUNTIF(Findings!C:C,"{check}")',
            desc,
            rating,
        ], 1):
            c = ws2.cell(i, col, val)
            c.fill      = rbg
            c.font      = _font(size=9)
            c.border    = _BORDER
            c.alignment = _align(h='center' if col == 2 else 'left', v='center')
        ws2.row_dimensions[i].height = 18

    # ── Sheet 3: Regelwerk ───────────────────────────────────────
    ws3 = wb.create_sheet('Regelwerk')
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions['A'].width = 46   # Check
    ws3.column_dimensions['B'].width = 42   # Regel
    ws3.column_dimensions['C'].width = 14   # Severity
    ws3.column_dimensions['D'].width = 70   # Erklärung

    # Banner
    ws3.merge_cells('A1:D1')
    b3 = ws3.cell(1, 1, 'DFIR Regelwerk — Stage 8.5 Check-Erläuterungen')
    b3.font      = _font(bold=True, color='FFFFFF', size=12)
    b3.fill      = F_BANNER
    b3.alignment = _align(h='center', v='center')
    ws3.row_dimensions[1].height = 24

    # Spalten-Header
    for col, lbl in enumerate(['Check', 'Regel', 'Severity', 'Erklärung'], 1):
        c = ws3.cell(2, col, lbl)
        c.font      = _font(bold=True, color='FFFFFF', size=10)
        c.fill      = F_HEADER
        c.alignment = _align(h='center', v='center')
        c.border    = _BORDER
    ws3.row_dimensions[2].height = 20
    ws3.freeze_panes = 'A3'

    REGELWERK = [
        (
            'M < B in /tmp/, /home/, /root/, /dev/shm/, /run/user/',
            'Timestamp-Manipulation in Verdachtspfad',
            'CRITICAL',
            'Die Änderungszeit (mtime) liegt vor dem Erstellzeitpunkt (birthtime) in einem '
            'typischen Angreifer-Pfad. Dies deutet auf gezieltes Timestomping hin — ein '
            'nachträgliches Zurücksetzen des Timestamps zur Verschleierung des Zugriffszeitpunkts.',
        ),
        (
            'M < B in /opt/, /srv/, /var/',
            'Möglicherweise rsync/tar oder Manipulation',
            'HIGH',
            'mtime liegt vor birthtime in einem neutralen Pfad, der weder rein dpkg-verwaltet '
            'noch klassischer Angreifer-Pfad ist. Kann durch Backup-Tools (rsync, tar --preserve) '
            'entstehen, erfordert aber Einzelfallprüfung.',
        ),
        (
            'C > M + 1h',
            'ctime viel neuer als mtime — touch -t Verdacht',
            'HIGH',
            'Die Change-Time (ctime) ist über eine Stunde neuer als die Modify-Time (mtime). '
            'Da ctime vom Kernel intern gesetzt und nicht direkt manipulierbar ist, '
            'weist eine solche Diskrepanz auf den Einsatz von touch -t zur mtime-Fälschung hin.',
        ),
        (
            'Timestamp vor Jahr 2000',
            'Unrealistisch auf modernem System',
            'HIGH',
            'Timestamps vor dem Jahr 2000 sind auf modernen Linux-Systemen strukturell '
            'ausgeschlossen. Solche Werte entstehen typischerweise durch manuelle Manipulation '
            '(touch -t 19990101) oder fehlerhafte Exploit-Tools.',
        ),
        (
            'Activity Burst (>10 Systemdateien in 5 Min)',
            'Rootkit-typisch',
            'HIGH',
            'Mehr als 10 Systemdateien wurden innerhalb von 5 Minuten modifiziert. Dieses '
            'Muster ist charakteristisch für Rootkit-Installationen, die viele Systembinaries '
            'gleichzeitig ersetzen oder patchen.',
        ),
        (
            'Staging Area (Datei in /tmp/, /dev/shm/ etc.)',
            'Angreifer-Ablagepfad',
            'MEDIUM',
            'Eine Datei befindet sich in einem temporären Verzeichnis, das häufig als '
            'Zwischenspeicher für Angreifer-Tools genutzt wird. Allein kein Beweis, '
            'aber im Kontext weiterer Findings ein starker Verdachtsindikator.',
        ),
        (
            'Deleted System File',
            'Gelöschter Inode in Systempfad',
            'HIGH',
            'Ein Inode in einem Systempfad ist noch vorhanden, aber nicht mehr im Verzeichnisbaum '
            'verlinkt (gelöschte Datei). Angreifer löschen Tools nach Ausführung — der Inode '
            'kann jedoch noch forensisch rekonstruiert werden.',
        ),
        (
            'Night Activity (00:00–05:00 Uhr)',
            'Systemaktivität zur Nachtzeit',
            'MEDIUM',
            'Dateisystemaktivität zwischen Mitternacht und 5 Uhr morgens. Kann legitime '
            'Cronjobs (Backup, Update) sein, ist in Kombination mit anderen Findings aber '
            'ein Hinweis auf automatisierten Angriff oder persistente Backdoor-Kommunikation.',
        ),
        (
            'Sorter Mismatch + detected = exec',
            'Dateitarnung als Executable',
            'CRITICAL',
            'Die Dateiendung suggeriert ein harmloses Format (z.B. .pdf, .jpg), Magic-Bytes '
            'belegen jedoch ein ELF-Binary oder Shell-Script. Klassische Tarnungstechnik, '
            'um Endpoint-Filter zu umgehen und ausführbaren Code zu verbergen.',
        ),
        (
            'Sorter Mismatch + detected ≠ exec',
            'Dateitarnung (andere Kategorie)',
            'HIGH',
            'Dateiendung und tatsächlicher Dateityp (laut Magic-Bytes) stimmen nicht überein, '
            'das Ergebnis ist aber kein ausführbares Binary. Kann Obfuskierung von Archiven, '
            'verschlüsselten Containern oder Dokumenten anzeigen.',
        ),
        (
            'Sorter Mismatch + MACtime-Anomalie auf gleicher Datei',
            'Doppelte Verschleierung',
            'CRITICAL',
            'Eine Datei zeigt gleichzeitig Typ-Fälschung (Sorter-Mismatch) und Timestamp-'
            'Manipulation (MACtime-Anomalie). Das gleichzeitige Auftreten beider Techniken '
            'auf derselben Datei ist ein starker Indikator für gezieltes Anti-Forensics.',
        ),
        (
            'Exec in Staging (laut Sorter)',
            'Executable in /tmp/ etc.',
            'CRITICAL',
            'Sorter hat eine ausführbare Datei (ELF-Binary, Shell-Script) in einem Staging-'
            'Verzeichnis identifiziert. Dies ist ein klassischer Dropper-Indikator: Angreifer '
            'laden Payloads in temporäre Verzeichnisse und starten sie von dort.',
        ),
        (
            'CVE-2021-4034 / CVE-2022-0847 + Systemdatei ±30 Min',
            'Bekannte kritische CVEs (PwnKit, Dirty Pipe)',
            'CRITICAL',
            'Systemdatei-Aktivität innerhalb von 30 Minuten um einen bekannten kritischen '
            'CVE-Exploit-Zeitpunkt. Die zeitliche Korrelation dieser Privilege-Escalation-CVEs '
            'mit Systemdatei-Modifikationen ist ein starker Kompromittierungshinweis.',
        ),
        (
            'Andere CVEs + Systemdatei ±30 Min',
            'CVE-Zeitfenster',
            'HIGH',
            'Systemdatei-Aktivität im ±30-Minuten-Fenster um einen anderen bekannten CVE-Exploit-'
            'Zeitpunkt. Der Verdachtsgrad ist geringer als bei kritischen CVEs, erfordert aber '
            'manuelle Untersuchung ob ein zeitlicher Zusammenhang besteht.',
        ),
        (
            'Stage-9-Upgrade (Timestomping bestätigt)',
            'Cross-Referenz Stage 9',
            '→ CRITICAL',
            'Stage 9 (Hayabusa/Sigma) hat unabhängig Timestomping-Aktivität auf derselben '
            'Datei detektiert. Die Bestätigung durch zwei unabhängige Analyse-Methoden '
            'erhöht die Konfidenz deutlich — das Finding wird automatisch auf CRITICAL hochgestuft.',
        ),
        (
            'Stage-9-Upgrade (Rootkit-Indikator bestätigt)',
            'Cross-Referenz Stage 9',
            '→ CRITICAL',
            'Stage 9 hat unabhängig einen Rootkit-Indikator auf derselben Datei identifiziert. '
            'Die Kombination aus Timeline-Anomalie (Stage 8.5) und Sigma-Regelübereinstimmung '
            '(Stage 9) ist ein starker Kompromittierungsnachweis.',
        ),
    ]

    for i, (check, regel, severity, erklaerung) in enumerate(REGELWERK, 3):
        rbg        = F_W if i % 2 == 1 else F_ALT
        sev_key    = severity.lstrip('→ ').strip()
        badge_fill = BADGE.get(sev_key, _fill('5F5E5A'))

        for col, val in enumerate([check, regel, severity, erklaerung], 1):
            c        = ws3.cell(i, col, val)
            c.border = _BORDER
            if col == 3:   # Severity-Badge
                c.fill      = badge_fill
                c.font      = _font(bold=True, color='FFFFFF', size=9)
                c.alignment = _align(h='center', v='center')
            else:
                c.fill      = rbg
                c.font      = _font(size=9)
                c.alignment = _align(h='left', v='top', wrap=True)
        ws3.row_dimensions[i].height = 38

    # ── Sheet 4: Sorter-Übersicht ────────────────────────────────
    sorter_files = getattr(ctx, 'tsk_sorter_files', {}) or {}

    # Dateiendung → erwartete Kategorie (identisch mit stage_timeline_analysis.py)
    _EXT_EXPECTED = {
        '.jpg': 'images',  '.jpeg': 'images', '.png': 'images',
        '.gif': 'images',  '.bmp': 'images',
        '.pdf': 'documents',
        '.txt': 'text',    '.log':  'text',   '.conf': 'text',  '.csv': 'text',
        '.zip': 'archive', '.tar':  'archive', '.gz': 'archive', '.bz2': 'archive',
        '.mp3': 'audio',   '.wav':  'audio',
        '.mp4': 'video',   '.avi':  'video',
        '.sh':  'exec',    '.py':   'exec',
    }

    ws4 = wb.create_sheet('Sorter-Übersicht')
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions['A'].width = 50   # Datei
    ws4.column_dimensions['B'].width = 20   # Erkannter_Typ
    ws4.column_dimensions['C'].width = 20   # Erwarteter_Typ
    ws4.column_dimensions['D'].width = 24   # Status

    # Banner
    ws4.merge_cells('A1:D1')
    b4 = ws4.cell(1, 1, f'DFIR Sorter-Übersicht — {len(sorter_files)} klassifizierte Dateien')
    b4.font      = _font(bold=True, color='FFFFFF', size=12)
    b4.fill      = F_BANNER
    b4.alignment = _align(h='center', v='center')
    ws4.row_dimensions[1].height = 24

    # Header
    for col, lbl in enumerate(['Datei', 'Erkannter_Typ', 'Erwarteter_Typ', 'Status'], 1):
        c = ws4.cell(2, col, lbl)
        c.font      = _font(bold=True, color='FFFFFF', size=10)
        c.fill      = F_HEADER
        c.alignment = _align(h='center', v='center')
        c.border    = _BORDER
    ws4.row_dimensions[2].height = 20
    ws4.freeze_panes = 'A3'

    if not sorter_files:
        ws4.merge_cells('A3:D3')
        msg = ws4.cell(3, 1, 'Kein Sorter-Output vorhanden — MACtime/Sorter möglicherweise deaktiviert.')
        msg.font      = _font(size=9, color='5F5E5A')
        msg.alignment = _align(h='center', v='center')
        msg.fill      = F_ALT
    else:
        # Sortierung: Mismatches zuerst, dann alphabetisch nach Dateiname
        def _sort_key(item):
            fname, detected = item
            ext      = Path(fname).suffix.lower()
            expected = _EXT_EXPECTED.get(ext)
            is_match = (expected is None) or (expected == detected)
            return (0 if not is_match else 1, fname.lower())

        sorted_sorter = sorted(sorter_files.items(), key=_sort_key)

        # Farben: Übereinstimmung → grün, Mismatch → rot, unbekannte Endung → neutral
        F_MATCH_A  = _fill('E8F5E9')   # hellgrün
        F_MATCH_B  = _fill('D4EDDA')   # hellgrün alternierend
        F_MISS_A   = _fill('FFF5F5')   # hellrot
        F_MISS_B   = _fill('FFEDED')   # hellrot alternierend
        F_NONE_A   = F_W               # weiß (unbekannte Endung)
        F_NONE_B   = F_ALT             # leicht grau

        match_cnt  = 0
        miss_cnt   = 0
        none_cnt   = 0

        for i, (fname, detected) in enumerate(sorted_sorter, 3):
            ext      = Path(fname).suffix.lower()
            expected = _EXT_EXPECTED.get(ext)

            if expected is None:
                status   = '—  Keine Referenz'
                rbg      = F_NONE_A if none_cnt % 2 == 0 else F_NONE_B
                none_cnt += 1
                exp_disp = '—'
            elif expected == detected:
                status    = '✅  Übereinstimmung'
                rbg       = F_MATCH_A if match_cnt % 2 == 0 else F_MATCH_B
                match_cnt += 1
                exp_disp  = expected
            else:
                status   = '⚠  Mismatch'
                rbg      = F_MISS_A if miss_cnt % 2 == 0 else F_MISS_B
                miss_cnt += 1
                exp_disp = expected

            for col, val in enumerate([fname, detected, exp_disp, status], 1):
                c        = ws4.cell(i, col, val)
                c.fill   = rbg
                c.font   = _font(size=9)
                c.border = _BORDER
                c.alignment = _align(h='left', v='center')
            ws4.row_dimensions[i].height = 18

        # Zusammenfassung unten
        summary_row = len(sorted_sorter) + 4
        ws4.merge_cells(f'A{summary_row}:D{summary_row}')
        summary_txt = (
            f'Gesamt: {len(sorter_files)}  |  '
            f'✅ Übereinstimmung: {match_cnt}  |  '
            f'⚠ Mismatch: {miss_cnt}  |  '
            f'— Keine Referenz: {none_cnt}'
        )
        sc = ws4.cell(summary_row, 1, summary_txt)
        sc.font      = _font(bold=True, size=9, color='1B3A5C')
        sc.fill      = F_ALT
        sc.alignment = _align(h='center', v='center')
        sc.border    = _BORDER

    out = case_dir / 'forensic_findings.xlsx'
    wb.save(str(out))
    log.info(f'  forensic_findings.xlsx → {out}  ({len(findings)} Befunde)')


# ── IP-Session-Analyse Excel ─────────────────────────────────────────────────

def _write_ip_sessions_excel(ctx: PipelineContext, case_dir: Path) -> None:
    """Aggregiert erfolgreiche Logins pro IP → ip_sessions.xlsx."""
    if not ctx.events_db_path or not ctx.events_db_path.exists():
        log.warning('  ip_sessions.xlsx: keine events.db vorhanden')
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning('openpyxl nicht installiert — ip_sessions.xlsx übersprungen')
        return
    from utils.event_store import EventStore

    # ── Style-Helfer (gleich wie forensic_findings) ───────────────────────────
    def _fill(h): return PatternFill(fill_type='solid', fgColor=h.lstrip('#'))
    def _font(bold=False, color='000000', size=9):
        return Font(name='Arial', bold=bold, color=color.lstrip('#'), size=size)
    def _align(h='left', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    _side   = Side(style='thin', color='C8D5E8')
    _BDR    = Border(left=_side, right=_side, top=_side, bottom=_side)
    F_BANNER = _fill('0D1B2A'); F_HEADER = _fill('1B3A5C')
    F_W = _fill('FFFFFF');     F_ALT    = _fill('F2F5F9')
    F_EXT  = _fill('FFF4DC');  F_ROOT   = _fill('FFF5F5'); F_INT = _fill('F0F7FF')

    # ── Aggregation: ip → Statistik ───────────────────────────────────────────
    def _src_cat(source: str) -> str:
        """Kategorisiert die Quelle fuer den Abgleich auth/journal/wtmp/wtmpdb."""
        s = (source or '').lower()
        if 'wtmpdb' in s:                        return 'wtmpdb'
        if 'wtmp' in s or 'utmp' in s or 'lastlog' in s: return 'wtmp'
        if 'auth' in s:                          return 'auth'
        if 'journal' in s or 'systemd' in s:     return 'journal'
        if 'syslog' in s or 'messages' in s:     return 'journal'
        return 'other'

    ip_data:    dict = {}
    detail_rows: list = []

    with EventStore(ctx.events_db_path) as store:
        for event in store.iter_events():
            login = _detect_login(event)
            if login is None:
                continue
            user, ip, method = login
            # Lokale/Konsolen-Logins (ohne IP) unter Sammelschluessel fuehren,
            # damit auch Nicht-SSH-Anmeldungen sichtbar sind (Betreuer-Wunsch)
            key = ip if ip else '(lokal/Konsole)'
            ts  = event.timestamp
            if key not in ip_data:
                ip_data[key] = {
                    'first': ts, 'last': ts, 'count': 0,
                    'users': set(), 'sources': set(), 'source_cats': set(),
                    'methods': set(),
                    'ip_type': _ip_type(key) if ip else 'Lokal',
                }
            d = ip_data[key]
            if ts < d['first']: d['first'] = ts
            if ts > d['last']:  d['last']  = ts
            d['count'] += 1
            if user: d['users'].add(user)
            d['sources'].add(event.source or '')
            d['source_cats'].add(_src_cat(event.source or ''))
            d['methods'].add(method)
            detail_rows.append({
                'ip': key, 'ts': ts, 'user': user, 'method': method,
                'event_type': event.event_type,
                'source': event.source or '',
                'message': (event.message or '')[:200],
            })

    if not ip_data:
        log.info('  ip_sessions.xlsx: keine erfolgreichen Logins mit IP — übersprungen')
        return

    _type_ord = {'Extern': 0, 'Intern': 1, 'Loopback': 2}
    sorted_ips  = sorted(ip_data.items(),
                         key=lambda x: (_type_ord.get(x[1]['ip_type'], 3), -x[1]['count']))
    extern_ips  = [(ip, d) for ip, d in sorted_ips if d['ip_type'] == 'Extern']
    detail_rows.sort(key=lambda r: (r['ip'], r['ts']))

    HDR  = ['IP', 'Typ', 'Login_Anzahl', 'Login_Methoden', 'Erster_Login_UTC',
            'Letzter_Login_UTC', 'Session_Dauer', 'Benutzer', 'Log-Quellen',
            'In_AuthLog', 'In_Journal', 'In_Wtmp', 'In_Wtmpdb']
    WID  = [16, 10, 13, 28, 21, 21, 14, 28, 28, 11, 11, 10, 11]
    LC   = get_column_letter(len(HDR))
    HDR3 = ['IP', 'Typ', 'Timestamp_UTC', 'Benutzer', 'Login_Methode',
            'Event_Type', 'Quelle', 'Nachricht']
    WID3 = [16, 10, 22, 16, 22, 22, 20, 58]
    LC3  = get_column_letter(len(HDR3))

    def _hdr_row(ws, row, headers, widths):
        for col, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row, col, h)
            c.font = _font(bold=True, color='FFFFFF', size=10)
            c.fill = F_HEADER; c.alignment = _align(h='center'); c.border = _BDR
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[row].height = 20

    def _banner(ws, row, text, lc):
        ws.merge_cells(f'A{row}:{lc}{row}')
        b = ws.cell(row, 1, text)
        b.font = _font(bold=True, color='FFFFFF', size=12)
        b.fill = F_BANNER; b.alignment = _align(h='center')
        ws.row_dimensions[row].height = 24

    def _ip_rows(ws, ip_list, start_row):
        for ri, (ip, d) in enumerate(ip_list, start_row):
            ipt    = d['ip_type']
            has_root = 'root' in d['users']
            rbg    = F_ROOT if (ipt == 'Extern' and has_root) \
                else F_EXT  if ipt == 'Extern' \
                else F_INT  if ipt == 'Intern' \
                else (F_W if ri % 2 == 0 else F_ALT)
            dur    = _fmt_duration(d['last'] - d['first']) \
                if d['first'] != d['last'] else '< 1 Min'
            scats  = d.get('source_cats', set())
            vals   = [ip, ipt, d['count'],
                      ', '.join(sorted(d.get('methods', set()))) or '—',
                      d['first'].strftime('%Y-%m-%d %H:%M:%S'),
                      d['last'].strftime('%Y-%m-%d %H:%M:%S'),
                      dur,
                      ', '.join(sorted(d['users'])) or '—',
                      ', '.join(sorted(d['sources'])) or '—',
                      '✓' if 'auth'    in scats else '✗',
                      '✓' if 'journal' in scats else '✗',
                      '✓' if 'wtmp'    in scats else '✗',
                      '✓' if 'wtmpdb'  in scats else '✗']
            for col, val in enumerate(vals, 1):
                c = ws.cell(ri, col, val)
                c.fill = rbg; c.border = _BDR
                c.font = _font(bold=(col == 1 and ipt == 'Extern' and has_root), size=9)
                c.alignment = _align(h='center' if col in (2, 3) else 'left',
                                     wrap=(col >= 8))
            ws.row_dimensions[ri].height = 18

    wb = Workbook()

    # Sheet 1: IP-Übersicht
    ws = wb.active; ws.title = 'IP-Übersicht'
    ws.sheet_view.showGridLines = False
    _banner(ws, 1,
            f'IP-Session-Analyse — {len(ip_data)} IPs — '
            f'{sum(d["count"] for _, d in sorted_ips)} erfolgreiche Logins', LC)
    _hdr_row(ws, 2, HDR, WID)
    ws.freeze_panes    = 'A3'
    ws.auto_filter.ref = f'A2:{LC}{len(sorted_ips) + 2}'
    _ip_rows(ws, sorted_ips, 3)

    # Sheet 2: Externe IPs
    ws2 = wb.create_sheet('Externe IPs')
    ws2.sheet_view.showGridLines = False
    _banner(ws2, 1, f'Externe IPs — {len(extern_ips)} IPs mit erfolgreichen Logins', LC)
    _hdr_row(ws2, 2, HDR, WID)
    ws2.freeze_panes = 'A3'
    if extern_ips:
        _ip_rows(ws2, extern_ips, 3)
    else:
        ws2.merge_cells(f'A3:{LC}3')
        nc = ws2.cell(3, 1, 'Keine externen IPs mit erfolgreichen Logins gefunden.')
        nc.font = _font(size=9, color='5F5E5A'); nc.alignment = _align(h='center')
        nc.fill = F_ALT

    # Sheet 3: Login-Detail
    ws3 = wb.create_sheet('Login-Detail')
    ws3.sheet_view.showGridLines = False
    _banner(ws3, 1, f'Login-Detail — {len(detail_rows)} Ereignisse', LC3)
    _hdr_row(ws3, 2, HDR3, WID3)
    ws3.freeze_panes    = 'A3'
    ws3.auto_filter.ref = f'A2:{LC3}{min(len(detail_rows) + 2, 65536)}'
    MAX_DETAIL = 5000
    for ri, row in enumerate(detail_rows[:MAX_DETAIL], 3):
        ipt = _ip_type(row['ip'])
        hr  = row['user'] == 'root'
        rbg = F_ROOT if (ipt == 'Extern' and hr) \
            else F_EXT if ipt == 'Extern' \
            else F_INT if ipt == 'Intern' \
            else (F_W if ri % 2 == 0 else F_ALT)
        for col, val in enumerate([row['ip'], ipt,
                                    row['ts'].strftime('%Y-%m-%d %H:%M:%S'),
                                    row['user'] or '—', row.get('method', '—'),
                                    row['event_type'],
                                    row['source'], row['message']], 1):
            c = ws3.cell(ri, col, val)
            c.fill = rbg; c.border = _BDR; c.font = _font(size=9)
            c.alignment = _align(h='left', wrap=(col == 8))
        ws3.row_dimensions[ri].height = 18
    if len(detail_rows) > MAX_DETAIL:
        nr = MAX_DETAIL + 3
        ws3.merge_cells(f'A{nr}:{LC3}{nr}')
        nc = ws3.cell(nr, 1,
                      f'... {len(detail_rows) - MAX_DETAIL} weitere Einträge '
                      f'abgeschnitten (Limit: {MAX_DETAIL})')
        nc.font = _font(size=9, color='5F5E5A'); nc.alignment = _align(h='center')

    # Sheet 4: Legende & Quellen-Abgleich (Betreuer-Wunsch: Provenienz + Limits)
    ws4 = wb.create_sheet('Legende')
    ws4.sheet_view.showGridLines = False
    ws4.column_dimensions['A'].width = 24
    ws4.column_dimensions['B'].width = 112
    _banner(ws4, 1, 'LEGENDE — Login-Quellen, Abgleich & Limitationen', 'B')
    _leg = [
        ('Datenquellen', 'Erfolgreiche Logins werden aus VIER Quellen zusammengefuehrt: '
            'auth.log/secure (SSH), systemd-Journal, wtmp und wtmpdb. '
            'Jede Zeile der Login-Detail-Mappe nennt in „Quelle" die konkrete Herkunft.'),
        ('In_AuthLog', '✓ = dieser Login ist in /var/log/auth.log (bzw. secure) belegt.'),
        ('In_Journal', '✓ = im systemd-Journal belegt (/var/log/journal/). '
            'Journale enthalten oft MEHR Logins als auth.log.'),
        ('In_Wtmp', '✓ = in /var/log/wtmp belegt (auch Konsolen-/lokale Logins).'),
        ('In_Wtmpdb', '✓ = in /var/lib/wtmpdb/wtmp.db belegt (moderner wtmp-Nachfolger).'),
        ('Abgleich', 'Stimmen die ✓/✗ ueber mehrere Quellen NICHT ueberein, kann das auf '
            'geloeschte/manipulierte Logs hindeuten (z.B. Login in wtmp, aber NICHT in '
            'auth.log). Genau dieser Quervergleich ist der forensische Mehrwert.'),
        ('Login_Methoden', 'SSH-Passwort, SSH-Schluessel, Remote (Netzwerk), '
            'Konsole/lokal, Session (Journal). Es werden NICHT nur SSH-Logins erfasst.'),
        ('(lokal/Konsole)', 'Sammelzeile fuer interaktive Logins OHNE Netzwerk-IP '
            '(direkte Konsole/tty, lokale Anmeldung).'),
        ('Limitationen', '(1) Journal-Logins werden ueber Nachrichtenmuster erkannt — '
            'untypische Formate koennen entgehen. (2) wtmp speichert den Remote-Host nur '
            'wenn vorhanden; lokale Logins erscheinen ohne IP. (3) Geloeschte Journal-/'
            'wtmp-Eintraege koennen nicht rekonstruiert werden. (4) Zeitzone aller '
            'Timestamps: UTC (in Stage 08 normalisiert).'),
    ]
    for ri, (k, v) in enumerate(_leg, 3):
        ws4.cell(ri, 1, k).font = _font(bold=True, size=9)
        c = ws4.cell(ri, 2, v); c.font = _font(size=9); c.alignment = _align(wrap=True)
        ws4.row_dimensions[ri].height = max(28, 14 * (len(v) // 95 + 1))

    out = case_dir / 'ip_sessions.xlsx'
    wb.save(str(out))
    log.info(f'  ip_sessions.xlsx → {out}  ({len(ip_data)} IPs/Quellen, {len(detail_rows)} Login-Events)')


# ── Reboot-Session-Analyse Excel ─────────────────────────────────────────────

def _write_reboot_sessions_excel(ctx: PipelineContext, case_dir: Path) -> None:
    """Paart Boot/Shutdown-Events zu Sessions → reboot_sessions.xlsx."""
    if not ctx.events_db_path or not ctx.events_db_path.exists():
        log.warning('  reboot_sessions.xlsx: keine events.db vorhanden')
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning('openpyxl nicht installiert — reboot_sessions.xlsx übersprungen')
        return
    from utils.event_store import EventStore

    def _fill(h): return PatternFill(fill_type='solid', fgColor=h.lstrip('#'))
    def _font(bold=False, color='000000', size=9):
        return Font(name='Arial', bold=bold, color=color.lstrip('#'), size=size)
    def _align(h='left', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    _side    = Side(style='thin', color='C8D5E8')
    _BDR     = Border(left=_side, right=_side, top=_side, bottom=_side)
    F_BANNER = _fill('0D1B2A'); F_HEADER = _fill('1B3A5C')
    F_W      = _fill('FFFFFF'); F_ALT    = _fill('F2F5F9')
    F_SHORT  = _fill('FFF4DC'); F_UNCLEAN= _fill('FFF5F5')
    _sev_badge = {
        'critical': _fill('A32D2D'), 'high':   _fill('BA7517'),
        'medium':   _fill('5F5E5A'), 'low':    _fill('3A6B3A'),
        'info':     _fill('3A6B3A'),
    }

    # Quell-Pfad-Mapping: Log-Typ → vollständiger Dateipfad auf dem Image
    _SRC_PATH_MAP = {
        'syslog':   '/var/log/syslog',
        'journal':  '/var/log/journal/ (systemd)',
        'auth':     '/var/log/auth.log',
        'auth.log': '/var/log/auth.log',
        'kern':     '/var/log/kern.log',
        'kern.log': '/var/log/kern.log',
        'daemon':   '/var/log/daemon.log',
        'boot':     '/var/log/boot.log',
        'messages': '/var/log/messages',
        'wtmp':     '/var/log/wtmp',
        'utmp':     '/var/run/utmp',
        'btmp':     '/var/log/btmp',
        'lastlog':  '/var/log/lastlog',
        'dpkg':     '/var/log/dpkg.log',
        'apt':      '/var/log/apt/history.log',
    }

    def _src_path_reboot(source: str) -> str:
        s = (source or '').lower().strip()
        return _SRC_PATH_MAP.get(s, source or '—')

    def _src_for_event(event) -> str:
        """Echte Quelle: Originalpfad aus der Extraktion (Provenienz-Feld).
        Im Fall-Modus mit [Image]-Praefix. Fallback: typischer Standardpfad."""
        op = getattr(event, 'orig_path', '') or ''
        base = op if op else _src_path_reboot(event.source)
        ev = getattr(event, 'evidence', '') or ''
        return f'[{ev}] {base}' if ev else base

    # ── Pass 1: Reboot-Events sammeln + Session-Pairing ───────────────────────
    reboot_events = []
    with EventStore(ctx.events_db_path) as store:
        for event in store.iter_events():
            if _classify_event(event) == 'REBOOT':
                kind = _classify_reboot_event(event)
                if kind:
                    reboot_events.append((event, kind))

    if not reboot_events:
        log.info('  reboot_sessions.xlsx: keine Boot/Shutdown-Events — übersprungen')
        return

    reboot_events.sort(key=lambda x: _to_utc(x[0].timestamp))

    # Pairing: boot → ... → shutdown
    sessions = []
    current_boot = None
    for event, kind in reboot_events:
        ts = event.timestamp
        if kind == 'boot':
            if current_boot is not None:
                sessions.append({'boot': current_boot, 'shutdown': None, 'unclean': True,
                                  'events': [], 'total': 0})
            current_boot = ts
        elif kind == 'shutdown':
            if current_boot is not None:
                sessions.append({'boot': current_boot, 'shutdown': ts, 'unclean': False,
                                  'events': [], 'total': 0})
                current_boot = None
            else:
                sessions.append({'boot': None, 'shutdown': ts, 'unclean': False,
                                  'events': [], 'total': 0})
    if current_boot is not None:
        sessions.append({'boot': current_boot, 'shutdown': None, 'unclean': False,
                          'events': [], 'total': 0})

    if not sessions:
        log.info('  reboot_sessions.xlsx: keine Sessions aufgebaut — übersprungen')
        return

    # ── Pass 2: Events den Sessions zuordnen (single pass) ────────────────────
    MAX_PER_SESSION = 2000
    with EventStore(ctx.events_db_path) as store:
        for event in store.iter_events():
            ts  = _to_utc(event.timestamp)
            sev = (getattr(event, 'severity', '') or '').lower()
            for sess in sessions:
                b = _to_utc(sess['boot'])     if sess['boot']     else None
                s = _to_utc(sess['shutdown']) if sess['shutdown'] else None
                if b and ts < b: continue
                if s and ts > s: continue
                sess['total'] += 1
                if sev in ('critical', 'high'):
                    sess['events'].append(event)
                elif len(sess['events']) < MAX_PER_SESSION:
                    sess['events'].append(event)
                break   # event gehört zur ersten passenden Session

    wb   = Workbook()
    HDR  = ['#', 'Boot_Start_UTC', 'Shutdown_UTC', 'Laufzeit', 'Ereignisse', 'Hinweis']
    WID  = [5, 22, 22, 16, 12, 50]
    LC   = get_column_letter(len(HDR))

    # ── Sheet 1: Übersicht ────────────────────────────────────────────────────
    ws = wb.active; ws.title = 'Übersicht'
    ws.sheet_view.showGridLines = False
    ws.merge_cells(f'A1:{LC}1')
    b = ws.cell(1, 1, f'Reboot-Session-Analyse — {len(sessions)} Betriebszeiträume')
    b.font = _font(bold=True, color='FFFFFF', size=12)
    b.fill = F_BANNER; b.alignment = _align(h='center')
    ws.row_dimensions[1].height = 24
    for col, (h, w) in enumerate(zip(HDR, WID), 1):
        c = ws.cell(2, col, h)
        c.font = _font(bold=True, color='FFFFFF', size=10)
        c.fill = F_HEADER; c.alignment = _align(h='center'); c.border = _BDR
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20
    ws.freeze_panes = 'A3'

    for ri, (i, sess) in enumerate(enumerate(sessions, 1), 3):
        bt = sess['boot']; st = sess['shutdown']
        unclean = sess.get('unclean', False)
        if bt and st:
            td      = _to_utc(st) - _to_utc(bt)
            dur     = _fmt_duration(td)
            is_short = td.total_seconds() < 1800   # < 30 Min
        elif bt:
            dur = 'Kein Shutdown'; is_short = False
        elif st:
            dur = 'Kein Boot-Beginn'; is_short = False
        else:
            dur = '?'; is_short = False

        if unclean:
            hint = '⚠ Ungeplanter Neustart (kein Shutdown vor nächstem Boot)'
        elif not st and bt:
            hint = 'Letzter bekannter Boot — kein sauberer Shutdown dokumentiert'
        elif not bt:
            hint = 'Shutdown vor erstem dokumentierten Boot'
        elif is_short:
            hint = f'⚠ Kurze Laufzeit ({dur}) — verdächtiger Neustart'
        else:
            hint = ''

        rbg = F_UNCLEAN if unclean else F_SHORT if is_short \
            else (F_W if ri % 2 == 0 else F_ALT)
        for col, val in enumerate([
            i,
            bt.strftime('%Y-%m-%d %H:%M:%S') if bt else '—',
            st.strftime('%Y-%m-%d %H:%M:%S') if st else '—',
            dur, sess['total'], hint,
        ], 1):
            c = ws.cell(ri, col, val)
            c.fill = rbg; c.border = _BDR
            c.font = _font(size=9, bold=(col == 1))
            c.alignment = _align(h='center' if col in (1, 4, 5) else 'left',
                                  wrap=(col == 6))
        ws.row_dimensions[ri].height = 20

    # ── Pro-Reboot-Sheets ─────────────────────────────────────────────────────
    MAX_SHEETS        = 20
    MAX_SHEET_ROWS    = 500
    HDR_D = ['Timestamp_UTC', 'Severity', 'Event_Type', 'Benutzer', 'IP', 'Quelle', 'Nachricht']
    WID_D = [22, 12, 24, 15, 16, 20, 70]
    LC_D  = get_column_letter(len(HDR_D))
    _sev_ord = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3, 'info': 4}

    for sess_nr, sess in enumerate(sessions[:MAX_SHEETS], 1):
        if not sess['events']:
            continue
        bt = sess['boot']; st = sess['shutdown']
        evs = sess['events']
        # HIGH/CRITICAL immer, dann Rest sortiert bis MAX_SHEET_ROWS
        if len(evs) > MAX_SHEET_ROWS:
            hi   = [e for e in evs if (getattr(e,'severity','') or '').lower() in ('critical','high')]
            rest = [e for e in evs if (getattr(e,'severity','') or '').lower() not in ('critical','high')]
            rest.sort(key=lambda e: (_sev_ord.get((getattr(e,'severity','') or '').lower(), 4),
                                     _to_utc(e.timestamp)))
            evs = hi + rest[:max(0, MAX_SHEET_ROWS - len(hi))]
        evs.sort(key=lambda e: _to_utc(e.timestamp))

        ws_r = wb.create_sheet(f'Reboot_{sess_nr}')
        ws_r.sheet_view.showGridLines = False
        bs = bt.strftime('%Y-%m-%d %H:%M') if bt else '?'
        ss = st.strftime('%Y-%m-%d %H:%M') if st else '?'
        ws_r.merge_cells(f'A1:{LC_D}1')
        br = ws_r.cell(1, 1,
                       f'Reboot {sess_nr}  Boot: {bs}  Shutdown: {ss}  '
                       f'({sess["total"]:,} Ereignisse)')
        br.font = _font(bold=True, color='FFFFFF', size=11)
        br.fill = F_BANNER; br.alignment = _align(h='center')
        ws_r.row_dimensions[1].height = 22
        for col, (h, w) in enumerate(zip(HDR_D, WID_D), 1):
            c = ws_r.cell(2, col, h)
            c.font = _font(bold=True, color='FFFFFF', size=10)
            c.fill = F_HEADER; c.alignment = _align(h='center'); c.border = _BDR
            ws_r.column_dimensions[get_column_letter(col)].width = w
        ws_r.row_dimensions[2].height = 20
        ws_r.freeze_panes    = 'A3'
        ws_r.auto_filter.ref = f'A2:{LC_D}{MAX_SHEET_ROWS + 3}'

        for ri, event in enumerate(evs, 3):
            sev   = (getattr(event, 'severity', 'info') or 'info').lower()
            badge = _sev_badge.get(sev, _fill('5F5E5A'))
            rbg   = F_W if ri % 2 == 0 else F_ALT
            for col, val in enumerate([
                event.timestamp.strftime('%Y-%m-%d %H:%M:%S') if event.timestamp else '',
                sev.upper(), event.event_type or '',
                event.user or '', event.ip or '',
                _src_for_event(event), event.message[:200],
            ], 1):
                c = ws_r.cell(ri, col, val)
                c.border = _BDR
                c.alignment = _align(h='center' if col == 2 else 'left', wrap=(col == 7))
                if col == 2:
                    c.fill = badge; c.font = _font(bold=True, color='FFFFFF', size=9)
                else:
                    c.fill = rbg; c.font = _font(size=9)
            ws_r.row_dimensions[ri].height = 18

        if sess['total'] > MAX_SHEET_ROWS:
            nr = len(evs) + 3
            ws_r.merge_cells(f'A{nr}:{LC_D}{nr}')
            nc = ws_r.cell(nr, 1,
                           f'Hinweis: {sess["total"] - len(evs):,} weitere Ereignisse '
                           f'nicht angezeigt (HIGH/CRITICAL priorisiert, Limit: {MAX_SHEET_ROWS})')
            nc.font = _font(size=9, color='5F5E5A'); nc.alignment = _align(h='center')

    # ── Legende-Sheet ─────────────────────────────────────────────────────────
    ws_leg = wb.create_sheet('Legende')
    ws_leg.sheet_view.showGridLines = False
    ws_leg.column_dimensions['A'].width = 28
    ws_leg.column_dimensions['B'].width = 80

    def _leg_banner(text, row):
        ws_leg.merge_cells(f'A{row}:B{row}')
        c = ws_leg.cell(row, 1, text)
        c.font      = _font(bold=True, color='FFFFFF', size=12)
        c.fill      = F_BANNER
        c.alignment = _align(h='center')
        ws_leg.row_dimensions[row].height = 24

    def _leg_header(row):
        for col, lbl in enumerate(['Bezeichnung', 'Erklärung'], 1):
            c = ws_leg.cell(row, col, lbl)
            c.font      = _font(bold=True, color='FFFFFF', size=10)
            c.fill      = F_HEADER
            c.alignment = _align(h='center')
            c.border    = _BDR
        ws_leg.row_dimensions[row].height = 20

    def _leg_rows(data, start_row):
        for i, (key, val) in enumerate(data, start_row):
            rbg = F_W if i % 2 == 0 else F_ALT
            for col, text in enumerate([key, val], 1):
                c        = ws_leg.cell(i, col, text)
                c.fill   = rbg
                c.font   = _font(bold=(col == 1), size=9)
                c.border = _BDR
                c.alignment = _align(h='left', wrap=True)
            ws_leg.row_dimensions[i].height = 32

    # Abschnitt 1 — Aufbau der Excel
    _leg_banner('Aufbau dieser Excel-Datei', 1)
    _leg_header(2)
    _leg_rows([
        ('Mappe: Übersicht',
         'Zeigt alle erkannten Boot/Shutdown-Zyklen als kompakte Tabelle. '
         'Jede Zeile = ein Betriebszeitraum (Boot → Shutdown). '
         'Ungeplante Neustarts (kein Shutdown vor nächstem Boot) sind orange markiert. '
         'Kurze Laufzeiten unter 30 Minuten sind gelb markiert — forensisch verdächtig.'),
        ('Mappe: Reboot_1, Reboot_2, ...',
         'Pro erkanntem Betriebszeitraum eine eigene Mappe mit allen Ereignissen '
         'die in diesem Zeitfenster aufgetreten sind. HIGH/CRITICAL-Ereignisse '
         'werden immer angezeigt, restliche Ereignisse bis max. 500 Zeilen. '
         'Spalten sind filterbar — Autofilter in Zeile 2.'),
    ], 3)

    # Abschnitt 2 — Woher kommen die Daten
    _leg_banner('Woher kommen die Daten?', 6)
    _leg_header(7)
    _leg_rows([
        ('Quelle allgemein',
         'Alle Ereignisse stammen aus den Log-Parsern der Pipeline (Stage 06). '
         'Die Logs wurden direkt aus dem Disk-Image extrahiert (TSK icat) '
         'und durch 38 spezialisierte Parser verarbeitet.'),
        ('Boot-Ereignisse',
         'Erkannt über Event-Types: boot, system_boot, kernel_start '
         'sowie Keywords: "system boot", "kernel command line", "linux version", '
         '"reached target basic system" — typischerweise aus /var/log/syslog '
         'oder /var/log/journal/.'),
        ('Shutdown-Ereignisse',
         'Erkannt über Event-Types: shutdown, system_shutdown '
         'sowie Keywords: "shutting down", "halt", "poweroff", '
         '"reached target shutdown" — typischerweise aus /var/log/syslog.'),
        ('Spalte "Quelle"',
         'Originalpfad der Log-Datei auf dem Image, dokumentiert bei der '
         'Extraktion (TSK fls/icat, siehe extraction_manifest.json). '
         'Direkt im Image nachpruefbar. Bei Quellen ohne Manifest-Eintrag '
         '(z.B. Hayabusa) wird der typische Standardpfad angezeigt.'),
    ], 8)

    # Abschnitt 3 — Spalten der Reboot-Sheets
    _leg_banner('Spalten der Reboot_X-Mappen', 13)
    _leg_header(14)
    _leg_rows([
        ('Timestamp_UTC',
         'Zeitstempel des Ereignisses in UTC. Normalisiert durch die Pipeline '
         '— Zeitzonenkorrektur wurde automatisch angewendet.'),
        ('Severity',
         'Schweregrad: CRITICAL (rot) → HIGH (orange) → MEDIUM (grau) → '
         'LOW/INFO (grün). Wird vom jeweiligen Parser oder der '
         'Timeline-Analyse vergeben.'),
        ('Event_Type',
         'Interner Ereignistyp des Parsers. Beispiele: boot, shutdown, '
         'kernel_start, system_boot — ermöglicht gezielte Filterung.'),
        ('Benutzer',
         'Benutzer der dem Ereignis zugeordnet ist, falls im Log vorhanden.'),
        ('IP',
         'IP-Adresse die dem Ereignis zugeordnet ist, falls im Log vorhanden.'),
        ('Quelle',
         'Originalpfad der Log-Quelldatei auf dem Image (aus dem '
         'Extraktions-Manifest). Direkt nachpruefbar — z.B. /var/log/syslog '
         'enthaelt die Boot-Kernel-Meldungen.'),
        ('Nachricht',
         'Original-Log-Zeile (gekürzt auf 200 Zeichen). '
         'Zeigt den genauen Wortlaut wie er in der Log-Datei steht.'),
    ], 15)

    # Abschnitt 4 — Farbkodierung
    _leg_banner('Farbkodierung der Übersicht', 23)
    _leg_header(24)
    _leg_rows([
        ('Rot/Rosa hinterlegt',
         'Ungeplanter Neustart — kein sauberer Shutdown vor dem nächsten Boot. '
         'Deutet auf Absturz, Stromausfall oder erzwungenen Neustart hin.'),
        ('Gelb/Orange hinterlegt',
         'Kurze Laufzeit unter 30 Minuten — forensisch auffällig. '
         'Kann auf einen gezielten Neustart nach einer Manipulation hinweisen.'),
        ('Weiß/Hellgrau',
         'Normaler Betriebszeitraum ohne besondere Auffälligkeiten.'),
    ], 25)

    out = case_dir / 'reboot_sessions.xlsx'
    wb.save(str(out))
    log.info(
        f'  reboot_sessions.xlsx → {out}  '
        f'({len(sessions)} Sessions, {sum(s["total"] for s in sessions)} Events gesamt)'
    )


# ── IOC-Excel ────────────────────────────────────────────────────────────────

def _write_iocs_excel(ctx: PipelineContext, case_dir: Path) -> None:
    """Alle IOCs als Excel — iocs.xlsx.

    Mappe 'IOCs':        oeffentliche Indikatoren (Typ-gefiltert via Autofilter)
    Mappe 'ip_private':  private/reservierte IPs separat (Review-Fix #18)
    Mappe 'Legende':     Spalten- und Typ-Erklaerung
    Jede Zeile traegt ihre Quelle (Parser/Herkunft) + Kontext-Schnipsel.
    """
    if not ctx.iocs:
        log.info('  iocs.xlsx: keine IOCs — übersprungen')
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning('openpyxl nicht installiert — iocs.xlsx übersprungen')
        return

    def _fill(h): return PatternFill(fill_type='solid', fgColor=h.lstrip('#'))
    def _font(bold=False, color='000000', size=9):
        return Font(name='Arial', bold=bold, color=color.lstrip('#'), size=size)
    def _align(h='left', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    _side = Side(style='thin', color='C8D5E8')
    _BDR  = Border(left=_side, right=_side, top=_side, bottom=_side)
    F_BANNER = _fill('0D1B2A'); F_HEADER = _fill('1B3A5C')
    F_W = _fill('FFFFFF'); F_ALT = _fill('F2F5F9')

    HDR = ['Typ', 'Wert', 'Quelle (Parser/Herkunft)', 'Kontext']
    WID = [16, 52, 24, 70]
    LC  = get_column_letter(len(HDR))

    pub  = [i for i in ctx.iocs if i.type != 'ip_private']
    priv = [i for i in ctx.iocs if i.type == 'ip_private']

    wb = Workbook()

    def _sheet(ws, titel, daten):
        ws.sheet_view.showGridLines = False
        ws.merge_cells(f'A1:{LC}1')
        b = ws.cell(1, 1, titel)
        b.font = _font(bold=True, color='FFFFFF', size=12)
        b.fill = F_BANNER; b.alignment = _align(h='center')
        ws.row_dimensions[1].height = 24
        for col, (h, w) in enumerate(zip(HDR, WID), 1):
            c = ws.cell(2, col, h)
            c.font = _font(bold=True, color='FFFFFF', size=10)
            c.fill = F_HEADER; c.alignment = _align(h='center'); c.border = _BDR
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[2].height = 20
        ws.freeze_panes = 'A3'
        ws.auto_filter.ref = f'A2:{LC}{len(daten) + 2}'
        for ri, ioc in enumerate(sorted(daten, key=lambda x: (x.type, x.value)), 3):
            rbg = F_W if ri % 2 == 0 else F_ALT
            for col, val in enumerate([ioc.type, ioc.value,
                                       ioc.source or '—',
                                       (ioc.context or '')[:200]], 1):
                c = ws.cell(ri, col, val)
                c.border = _BDR; c.fill = rbg
                c.font = _font(size=9)
                c.alignment = _align(wrap=(col == 4))

    ws1 = wb.active; ws1.title = 'IOCs'
    _sheet(ws1, f'INDIKATOREN (IOCs) — {len(pub)} Eintraege', pub)
    if priv:
        ws2 = wb.create_sheet('ip_private')
        _sheet(ws2, f'PRIVATE / RESERVIERTE IPs — {len(priv)} Eintraege '
                    f'(separat: verrauschen oeffentliche IOCs nicht, '
                    f'Lateral Movement bleibt sichtbar)', priv)

    ws3 = wb.create_sheet('Legende')
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions['A'].width = 26
    ws3.column_dimensions['B'].width = 110
    legende = [
        ('Typ', 'IOC-Kategorie: ip, ip_private, ipv6, domain, url, hash_md5, '
                'hash_sha1, hash_sha256, email, cve, registry_key'),
        ('Wert', 'Der Indikator selbst — Hashes immer vollstaendig, keine Kuerzung.'),
        ('Quelle (Parser/Herkunft)', 'Welcher Parser bzw. welches Werkzeug den IOC fand '
                '(z.B. auth, bash_history, bulk_extractor). Rueckverfolgbar bis zur '
                'Ursprungsdatei ueber die Timeline (orig_path).'),
        ('Kontext', 'Textausschnitt rund um den Fund (±40 Zeichen).'),
        ('Mappe ip_private', 'Private/reservierte IPs (RFC1918, Loopback) separat — '
                'im LAN-Szenario relevant, verrauschen aber die oeffentlichen IOCs.'),
    ]
    for ri, (k, v) in enumerate(legende, 2):
        ws3.cell(ri, 1, k).font = _font(bold=True, size=9)
        c = ws3.cell(ri, 2, v); c.font = _font(size=9); c.alignment = _align(wrap=True)
        ws3.row_dimensions[ri].height = 28

    out = case_dir / 'iocs.xlsx'
    wb.save(str(out))
    log.info(f'  iocs.xlsx → {out}  ({len(pub)} IOCs, {len(priv)} ip_private)')


# ── Gefilterte Filesystem-Timeline Excel ─────────────────────────────────────

def _write_filtered_filesystem_timeline_excel(ctx: PipelineContext, case_dir: Path) -> None:
    """Gefilterte MACtime-Timeline (nur relevante Pfade) → filtered_filesystem_timeline.xlsx."""
    if not ctx.events_db_path or not ctx.events_db_path.exists():
        log.warning('  filtered_filesystem_timeline.xlsx: keine events.db vorhanden')
        return
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        log.warning('openpyxl nicht installiert — filtered_filesystem_timeline.xlsx übersprungen')
        return
    from utils.event_store import EventStore
    from collections import defaultdict

    def _fill(h): return PatternFill(fill_type='solid', fgColor=h.lstrip('#'))
    def _font(bold=False, color='000000', size=9):
        return Font(name='Arial', bold=bold, color=color.lstrip('#'), size=size)
    def _align(h='left', v='center', wrap=False):
        return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
    _side    = Side(style='thin', color='C8D5E8')
    _BDR     = Border(left=_side, right=_side, top=_side, bottom=_side)
    F_BANNER = _fill('0D1B2A'); F_HEADER = _fill('1B3A5C')
    _sev_badge = {
        'critical': _fill('A32D2D'), 'high':   _fill('BA7517'),
        'medium':   _fill('5F5E5A'), 'low':    _fill('3A6B3A'),
        'info':     _fill('3A6B3A'),
    }

    # Cross-Reference: Dateipfade aus forensic_findings
    finding_paths = {f.file for f in (ctx.forensic_findings or []) if f.file}

    filtered = []
    with EventStore(ctx.events_db_path) as store:
        for event in store.iter_events():
            if event.source != 'mactime':
                continue
            fp  = event.file_path or ''
            sev = (event.severity or 'info').lower()
            # Immer einschließen: HIGH/CRITICAL oder in forensic_findings
            if sev in ('high', 'critical') or fp in finding_paths:
                filtered.append(event)
                continue
            # Whitelist
            if not any(fp.startswith(p) for p in _INTERESTING_PREFIXES):
                continue
            # Blacklist
            if any(fp.startswith(p) for p in _NOISE_PREFIXES):
                continue
            filtered.append(event)

    if not filtered:
        log.info('  filtered_filesystem_timeline.xlsx: keine relevanten MACtime-Events — übersprungen')
        return

    filtered.sort(key=lambda e: _to_utc(e.timestamp))

    # Verzeichnis-Statistik
    dir_stats: dict = defaultdict(lambda: {'count': 0, 'first': None, 'last': None})
    for event in filtered:
        from pathlib import PurePosixPath
        fp = event.file_path or ''
        d  = str(PurePosixPath(fp).parent) if fp else '?'
        ds = dir_stats[d]
        ds['count'] += 1
        ts = _to_utc(event.timestamp)
        if ds['first'] is None or ts < ds['first']: ds['first'] = ts
        if ds['last']  is None or ts > ds['last']:  ds['last']  = ts
    dir_sorted = sorted(dir_stats.items(), key=lambda x: -x[1]['count'])

    # ── Gruppierung: eine Zeile pro Datei mit allen 4 Timestamps ────────────
    from collections import defaultdict as _dd
    _SEV_ORD_TL = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3, 'info': 4}

    file_ts: dict = _dd(lambda: {
        'm': None, 'a': None, 'c': None, 'b': None,
        'severity': 'info', 'category': 'Sonstige',
        'fill_a': 'FFFFFF', 'fill_b': 'F2F5F9',
    })

    for event in filtered:
        fp  = (event.file_path or '').strip()
        et  = (event.event_type or '').lower().replace('filesystem_', '')
        sev = (event.severity or 'info').lower()
        ts  = _to_utc(event.timestamp) if event.timestamp else None

        for letter in ('m', 'a', 'c', 'b'):
            if letter in et and ts is not None:
                existing = file_ts[fp][letter]
                if existing is None or ts > existing:
                    file_ts[fp][letter] = ts

        if _SEV_ORD_TL.get(sev, 4) < _SEV_ORD_TL.get(file_ts[fp]['severity'], 4):
            file_ts[fp]['severity'] = sev

        cat, fa, fb = _path_category(fp)
        file_ts[fp]['category'] = cat
        file_ts[fp]['fill_a']   = fa
        file_ts[fp]['fill_b']   = fb

    def _fmt_ts_tl(ts):
        return ts.strftime('%Y-%m-%d %H:%M:%S') if ts else '—'

    def _auffaelligkeit(m, a, c, b):
        hints = []
        if m and c:
            diff = (_to_utc(c) - _to_utc(m)).total_seconds()
            if diff > 3600:
                hints.append('⚠ ctime > mtime+1h → Timestomping?')
        if b and m and _to_utc(b) > _to_utc(m):
            hints.append('⚠ btime > mtime → Datei kopiert?')
        if m and _to_utc(m).year < 2000:
            hints.append('🔴 mtime vor 2000 → Verdächtig')
        if b and _to_utc(b).year < 2000:
            hints.append('🔴 btime vor 2000 → Verdächtig')
        return ' | '.join(hints) if hints else '—'

    file_rows = sorted(
        [{'fp': fp, **data} for fp, data in file_ts.items()],
        key=lambda r: (_SEV_ORD_TL.get(r['severity'], 4), r['fp'])
    )

    wb  = Workbook()
    HDR = ['Datei_Pfad', 'mtime_UTC', 'atime_UTC', 'ctime_UTC', 'btime_UTC',
           'Auffälligkeit', 'Severity', 'Kategorie']
    WID = [60, 22, 22, 22, 22, 42, 12, 14]
    LC  = get_column_letter(len(HDR))

    # ── Sheet 1: Timeline ─────────────────────────────────────────────────────
    ws = wb.active; ws.title = 'Timeline'
    ws.sheet_view.showGridLines = False
    ws.merge_cells(f'A1:{LC}1')
    b = ws.cell(1, 1,
                f'Gefilterte Filesystem-Timeline — {len(file_rows):,} Dateien '
                f'({len(filtered):,} MACB-Events gruppiert) — '
                f'Quelle: TSK fls -m → mactime (Inode-Metadaten direkt aus Disk-Image)')
    b.font = _font(bold=True, color='FFFFFF', size=12)
    b.fill = F_BANNER; b.alignment = _align(h='center')
    ws.row_dimensions[1].height = 24
    for col, (h, w) in enumerate(zip(HDR, WID), 1):
        c = ws.cell(2, col, h)
        c.font = _font(bold=True, color='FFFFFF', size=10)
        c.fill = F_HEADER; c.alignment = _align(h='center'); c.border = _BDR
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[2].height = 20
    ws.freeze_panes    = 'A3'
    ws.auto_filter.ref = f'A2:{LC}{len(file_rows) + 2}'

    for ri, row in enumerate(file_rows, 3):
        sev = row['severity']
        rbg = _fill(row['fill_a']) if ri % 2 == 0 else _fill(row['fill_b'])
        auf = _auffaelligkeit(row['m'], row['a'], row['c'], row['b'])
        for col, val in enumerate([
            row['fp'],
            _fmt_ts_tl(row['m']),
            _fmt_ts_tl(row['a']),
            _fmt_ts_tl(row['c']),
            _fmt_ts_tl(row['b']),
            auf,
            sev.upper(),
            row['category'],
        ], 1):
            c = ws.cell(ri, col, val)
            c.border = _BDR
            c.alignment = _align(
                h='center' if col in (7, 8) else 'left',
                wrap=(col == 6),
            )
            if col == 7:
                c.fill = _sev_badge.get(sev, _fill('5F5E5A'))
                c.font = _font(bold=True, color='FFFFFF', size=9)
            elif col == 6 and auf != '—':
                c.fill = _fill('FFF5F5') if ri % 2 == 0 else _fill('FFEDED')
                c.font = _font(size=9, color='A32D2D')
            else:
                c.fill = rbg; c.font = _font(size=9)
        ws.row_dimensions[ri].height = 20

    # ── Sheet 2: Verzeichnis-Übersicht ────────────────────────────────────────
    ws2 = wb.create_sheet('Verzeichnis-Übersicht')
    ws2.sheet_view.showGridLines = False
    for col, w in enumerate([65, 14, 22, 22, 14], 1):
        ws2.column_dimensions[get_column_letter(col)].width = w
    ws2.merge_cells('A1:E1')
    b2 = ws2.cell(1, 1,
                  f'Verzeichnis-Übersicht — {len(dir_sorted)} Verzeichnisse '
                  f'mit relevanten MACtime-Events')
    b2.font = _font(bold=True, color='FFFFFF', size=12)
    b2.fill = F_BANNER; b2.alignment = _align(h='center')
    ws2.row_dimensions[1].height = 24
    for col, lbl in enumerate(
            ['Verzeichnis', 'Anzahl_Events', 'Ersten_Timestamp', 'Letzten_Timestamp', 'Kategorie'],
            1):
        c = ws2.cell(2, col, lbl)
        c.font = _font(bold=True, color='FFFFFF', size=10)
        c.fill = F_HEADER; c.alignment = _align(h='center'); c.border = _BDR
    ws2.row_dimensions[2].height = 20
    ws2.freeze_panes = 'A3'

    for ri, (dirpath, ds) in enumerate(dir_sorted, 3):
        cat, fa, _ = _path_category(dirpath + '/')
        rbg        = _fill(fa)
        for col, val in enumerate([
            dirpath, ds['count'],
            ds['first'].strftime('%Y-%m-%d %H:%M:%S') if ds['first'] else '',
            ds['last'].strftime('%Y-%m-%d %H:%M:%S')  if ds['last']  else '',
            cat,
        ], 1):
            c = ws2.cell(ri, col, val)
            c.fill = rbg; c.border = _BDR; c.font = _font(size=9)
            c.alignment = _align(h='center' if col == 2 else 'left')
        ws2.row_dimensions[ri].height = 18

    # ── Sheet 3: Legende ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet('Legende')
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions['A'].width = 10   # Kürzel
    ws3.column_dimensions['B'].width = 10   # Name
    ws3.column_dimensions['C'].width = 62   # Bedeutung
    ws3.column_dimensions['D'].width = 36   # Manipulierbar

    # Banner
    ws3.merge_cells('A1:D1')
    b3 = ws3.cell(1, 1, 'Legende — MACB-Timestamps & Filterkriterien')
    b3.font      = _font(bold=True, color='FFFFFF', size=12)
    b3.fill      = F_BANNER
    b3.alignment = _align(h='center')
    ws3.row_dimensions[1].height = 24

    # Abschnitt 1: MACB-Tabelle
    ws3.merge_cells('A2:D2')
    s1 = ws3.cell(2, 1, 'MACB-Timestamp-Typen')
    s1.font      = _font(bold=True, color='FFFFFF', size=10)
    s1.fill      = F_HEADER
    s1.alignment = _align(h='left')
    ws3.row_dimensions[2].height = 20

    # Header
    F_W_leg  = _fill('FFFFFF')
    F_ALT_leg = _fill('F2F5F9')
    for col, lbl in enumerate(['Kürzel', 'Name', 'Bedeutung', 'Manipulierbar?'], 1):
        c = ws3.cell(3, col, lbl)
        c.font      = _font(bold=True, color='FFFFFF', size=10)
        c.fill      = F_HEADER
        c.alignment = _align(h='center')
        c.border    = _BDR
    ws3.row_dimensions[3].height = 20

    MACB_LEGEND = [
        ('m', 'mtime',
         'Wann wurde der Dateiinhalt zuletzt geändert?',
         '✅ Ja — mit touch -t manipulierbar'),
        ('a', 'atime',
         'Wann wurde die Datei zuletzt gelesen / geöffnet?',
         '✅ Ja — mit touch -a manipulierbar'),
        ('c', 'ctime',
         'Wann wurden die Metadaten zuletzt geändert (Rechte, Eigentümer, Links)?',
         '❌ Nein — wird nur vom Kernel gesetzt, nicht direkt manipulierbar'),
        ('b', 'btime',
         'Wann wurde die Datei erstellt (Birth/Creation Time)?',
         '✅ Eingeschränkt — nicht auf allen Dateisystemen verfügbar'),
    ]

    for i, (kurzel, name, bedeutung, manip) in enumerate(MACB_LEGEND, 4):
        rbg = F_W_leg if i % 2 == 0 else F_ALT_leg
        for col, val in enumerate([kurzel, name, bedeutung, manip], 1):
            c        = ws3.cell(i, col, val)
            c.fill   = rbg
            c.font   = _font(bold=(col == 1), size=9)
            c.border = _BDR
            c.alignment = _align(h='center' if col <= 2 else 'left', wrap=(col >= 3))
        ws3.row_dimensions[i].height = 28

    # Abschnitt 2: Forensische Bedeutung
    ws3.merge_cells('A9:D9')
    s2 = ws3.cell(9, 1, 'Forensische Bedeutung von Timestamp-Kombinationen')
    s2.font      = _font(bold=True, color='FFFFFF', size=10)
    s2.fill      = F_HEADER
    s2.alignment = _align(h='left')
    ws3.row_dimensions[9].height = 20

    FORENSIC_HINTS = [
        ('ctime > mtime + 1h',
         'Verdacht auf Timestomping — mtime wurde nachträglich mit touch -t zurückgesetzt. '
         'ctime kann nicht manuell manipuliert werden und verrät den echten Änderungszeitpunkt.'),
        ('btime viel neuer als atime',
         'Datei wurde wahrscheinlich kopiert, nicht neu erstellt. '
         'Beim Kopieren wird btime neu gesetzt, atime bleibt vom Original erhalten.'),
        ('mtime vor Installationsdatum des Systems',
         'Unrealistischer Timestamp — deutet auf manuelle Manipulation (Timestomping) hin.'),
        ('Timestamp vor Jahr 2000',
         'Auf modernen Linux-Systemen strukturell ausgeschlossen. '
         'Entsteht durch touch -t 19990101 oder fehlerhafte Exploit-Tools.'),
    ]

    for col, lbl in enumerate(['Kombination', 'Bedeutung'], 1):
        c = ws3.cell(10, col, lbl)
        c.font      = _font(bold=True, color='FFFFFF', size=10)
        c.fill      = F_HEADER
        c.alignment = _align(h='center')
        c.border    = _BDR
    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 82
    ws3.row_dimensions[10].height = 20

    for i, (kombi, bedeutung) in enumerate(FORENSIC_HINTS, 11):
        rbg = F_W_leg if i % 2 == 0 else F_ALT_leg
        for col, val in enumerate([kombi, bedeutung], 1):
            c        = ws3.cell(i, col, val)
            c.fill   = rbg
            c.font   = _font(size=9)
            c.border = _BDR
            c.alignment = _align(h='left', wrap=True)
        ws3.row_dimensions[i].height = 32

    # Abschnitt 3: Filterkriterien
    ws3.merge_cells('A16:D16')
    s3 = ws3.cell(16, 1, 'Warum sind genau diese Einträge in dieser Tabelle?')
    s3.font      = _font(bold=True, color='FFFFFF', size=10)
    s3.fill      = F_HEADER
    s3.alignment = _align(h='left')
    ws3.row_dimensions[16].height = 20

    FILTER_INFO = [
        ('Quelle der Timestamps',
         'TSK fls -m liest die rohen Inode-Metadaten direkt aus dem Disk-Image. '
         'mactime konvertiert das Body-File in eine lesbare Timeline. '
         'Die Timestamps stammen also direkt aus dem Dateisystem — nicht aus Logs.'),
        ('Immer enthalten',
         'Einträge mit Severity HIGH oder CRITICAL sowie Dateipfade die bereits '
         'in den forensic_findings auftauchen.'),
        ('Relevante Pfade (Whitelist)',
         '/home/  /root/  /tmp/  /var/tmp/  /dev/shm/  /etc/  '
         '/usr/local/  /opt/  /srv/  /var/www/  /run/  /var/spool/'),
        ('Herausgefiltert (Blacklist)',
         '/usr/lib/  /usr/share/  /lib/  /lib64/  '
         '/proc/  /sys/  /dev/  /run/lock/  /snap/  '
         '— System-Libraries ohne forensische Relevanz'),
    ]

    for col, lbl in enumerate(['Kriterium', 'Erklärung'], 1):
        c = ws3.cell(17, col, lbl)
        c.font      = _font(bold=True, color='FFFFFF', size=10)
        c.fill      = F_HEADER
        c.alignment = _align(h='center')
        c.border    = _BDR
    ws3.row_dimensions[17].height = 20

    for i, (kriterium, erklaerung) in enumerate(FILTER_INFO, 18):
        rbg = F_W_leg if i % 2 == 0 else F_ALT_leg
        for col, val in enumerate([kriterium, erklaerung], 1):
            c        = ws3.cell(i, col, val)
            c.fill   = rbg
            c.font   = _font(size=9)
            c.border = _BDR
            c.alignment = _align(h='left', wrap=True)
        ws3.row_dimensions[i].height = 36

    out = case_dir / 'filtered_filesystem_timeline.xlsx'
    wb.save(str(out))
    log.info(
        f'  filtered_filesystem_timeline.xlsx → {out}  '
        f'({len(filtered):,} gefilterte Events, {len(dir_sorted)} Verzeichnisse)'
    )


# ── PDF Report ───────────────────────────────────────────────────────────────

def _generate_report_pdf(ctx: PipelineContext, case_dir: Path) -> None:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.pdfgen import canvas
        from reportlab.lib import colors
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, PageBreak)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    except ImportError:
        log.warning('reportlab nicht installiert — PDF-Export übersprungen')
        return

    out_file = case_dir / 'report.pdf'
    doc = SimpleDocTemplate(str(out_file), pagesize=A4,
                            leftMargin=15*mm, rightMargin=15*mm,
                            topMargin=35*mm, bottomMargin=22*mm)

    styles    = getSampleStyleSheet()
    story     = []
    case_id   = case_dir.name
    created   = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')
    quality   = evaluate_quality(ctx)
    duration  = _safe_duration(ctx.start_time)

    def _rl_color(t):
        from reportlab.lib.colors import Color
        return Color(*t)

    def _h1(text):
        """Seitenüberschrift — dunkelblaue Unterlinie, gut sichtbar."""
        from reportlab.platypus import Table as _T, TableStyle as _TS
        p = Paragraph(
            f'<font size="15"><b><font color="#{_hex(C_DARK_BLUE)}">{text}</font></b></font>',
            styles['Normal']
        )
        t = _T([[p]], colWidths=[W])
        t.setStyle(_TS([
            ('LINEBELOW',     (0,0), (-1,-1), 2.5, _rl_color(C_DARK_BLUE)),
            ('TOPPADDING',    (0,0), (-1,-1), 2),
            ('BOTTOMPADDING', (0,0), (-1,-1), 7),
            ('LEFTPADDING',   (0,0), (-1,-1), 0),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
        ]))
        return t

    def _h2(text):
        """Abschnittsüberschrift — blauer Balken links, klar abgesetzt."""
        from reportlab.platypus import Table as _T, TableStyle as _TS
        p = Paragraph(
            f'<font size="10"><b><font color="#{_hex(C_MID_BLUE)}">{text}</font></b></font>',
            styles['Normal']
        )
        t = _T([['', p]], colWidths=[4*mm, W - 4*mm])
        t.setStyle(_TS([
            ('BACKGROUND',    (0,0), (0,-1), _rl_color(C_MID_BLUE)),
            ('TOPPADDING',    (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING',   (0,0), (0,-1), 0),
            ('LEFTPADDING',   (1,0), (1,-1), 8),
            ('RIGHTPADDING',  (0,0), (-1,-1), 0),
        ]))
        return t

    def _body(text):
        return Paragraph(f'<font size="9" color="#{_hex(C_DARK_GREY)}">{text}</font>',
                         styles['Normal'])

    def _spacer(h=4):
        return Spacer(1, h*mm)

    def _table(data, col_widths=None, header=True):
        """Datentabelle mit automatischem Zeilenumbruch in allen Zellen."""
        from reportlab.platypus import Table, TableStyle
        from reportlab.lib.styles import ParagraphStyle
        _hs = ParagraphStyle('th', fontSize=8, fontName='Helvetica-Bold',
                              textColor=_rl_color(C_WHITE), leading=11)
        _cs = ParagraphStyle('td', fontSize=8, fontName='Helvetica',
                              textColor=_rl_color(C_DARK_GREY), leading=11)
        wrapped = []
        for i, row in enumerate(data):
            wr = []
            for cell in row:
                if isinstance(cell, str):
                    st = _hs if (header and i == 0) else _cs
                    wr.append(Paragraph(cell.replace('\n', '<br/>'), st))
                else:
                    wr.append(cell)
            wrapped.append(wr)
        t = Table(wrapped, colWidths=col_widths, repeatRows=1 if header else 0)
        t.setStyle(TableStyle([
            ('BACKGROUND',    (0,0), (-1,0),  _rl_color(C_DARK_BLUE)),
            ('ROWBACKGROUNDS',(0,1), (-1,-1),
             [_rl_color(C_WHITE), _rl_color(C_LIGHT_GREY)]),
            ('GRID',          (0,0), (-1,-1), 0.3, _rl_color((0xDD/255,)*3)),
            ('LEFTPADDING',   (0,0), (-1,-1), 6),
            ('RIGHTPADDING',  (0,0), (-1,-1), 6),
            ('TOPPADDING',    (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('VALIGN',        (0,0), (-1,-1), 'TOP'),
        ]))
        return t

    W = A4[0] - 30*mm  # Nutzbare Breite

    # ── Seite 1: Deckblatt ───────────────────────────────────────────────────
    story.append(_spacer(5))
    story.append(_banner('DFIR ANALYSE-REPORT', C_DARK_BLUE, C_WHITE, 28, W))
    story.append(_spacer(2))
    story.append(_body('Automatisch generierter forensischer Analysebericht'))
    story.append(_spacer(8))

    # Review-Fix #19/#5: Hashes IMMER vollstaendig, Label nach Hash-Art
    # (E01 bettet MD5+SHA1 ein — frueher wurde SHA1 als 'SHA256' angezeigt
    #  und zusaetzlich auf 32 Zeichen gekuerzt)
    _e01 = ctx.hash_source == 'E01-eingebettet'
    kv1 = [
        ['Case-ID',    case_id],       ['Erstellt',    created],
        ['Dauer',      f'{duration} Minuten'], ['Format', ctx.file_type],
    ]
    if getattr(ctx, 'sha1', ''):
        kv1.append(['SHA1 (E01-eingebettet)', ctx.sha1])
    if ctx.sha256:
        kv1.append(['SHA256', ctx.sha256])
    kv1 += [
        ['MD5 (E01-eingebettet)' if _e01 else 'MD5', ctx.md5 or '?'],
        ['Qualitaet',  quality],       ['IOC-Qualitaet', ctx.ioc_quality],
    ]
    story.append(_h2('Case-Informationen'))
    story.append(_spacer(2))
    story.append(_kv_table(kv1, W, _rl_color))
    story.append(_spacer(6))

    # ── Fall-Modus: Uebersicht aller analysierten Images ──────────────────
    _evi = getattr(ctx, 'evidence_items', []) or []
    if getattr(ctx, 'combined_case', False) and _evi:
        story.append(_h2(f'Fall-Uebersicht — {len(_evi)} Beweisstuecke (Images)'))
        story.append(_spacer(2))
        ev_rows = [['#', 'Image', 'OS', 'Hostname', 'Format', 'Groesse']]
        for i, ev in enumerate(_evi, 1):
            ev_rows.append([
                str(i), ev.get('name', '-'),
                (ev.get('os_name') or ev.get('os_family') or '-')[:28],
                ev.get('hostname') or '-',
                ev.get('file_type', '-'),
                f"{ev.get('file_size_gb', 0):.1f} GB",
            ])
        story.append(_table(ev_rows, [10*mm, 42*mm, 46*mm, 32*mm, 18*mm, 22*mm]))
        story.append(_spacer(3))
        # Hashes pro Image (forensisch getrennt) — vollstaendig
        coc_ev = ctx.coc.evidence_hashes if ctx.coc else {}
        for ev in _evi:
            h = coc_ev.get(ev.get('name', ''), {})
            hr = [['Image', ev.get('name', '-')]]
            if h.get('sha1'):   hr.append(['SHA1 (E01)', h['sha1']])
            if h.get('sha256'): hr.append(['SHA256', h['sha256']])
            if h.get('md5'):    hr.append(['MD5', h['md5']])
            if len(hr) > 1:
                story.append(_kv_table(hr, W, _rl_color))
                story.append(_spacer(2))
        story.append(_spacer(4))

    kv2 = [
        ['OS',           ctx.os_name or '-'],
        ['Kernel',       ctx.kernel_version or '-'],
        ['Hostname',     ctx.hostname or '-'],
        ['Zeitzone',     getattr(ctx, 'timezone_display', None) or ctx.timezone or '-'],
        ['Autopsy',      'Gelaufen' if ctx.autopsy_ran else ctx.autopsy_reason[:60]],
        ['TSK-Fallback', 'Aktiv' if ctx.tsk_fallback_used else 'Nicht noetig'],
    ]
    story.append(_h2('System-Information'))
    story.append(_spacer(2))
    story.append(_kv_table(kv2, W, _rl_color))
    story.append(PageBreak())

    # ── Seite 2: Partition-Layout ────────────────────────────────────────────
    story.append(_h1('Partition-Layout'))
    story.append(_spacer(3))
    story.append(_body(
        'Uebersicht aller erkannten Partitionen im Disk-Image. '
        'Primaerpartition = Betriebssystem-Partition (OS-Erkennung). '
        'Offset in 512-Byte-Sektoren (TSK-Standard).'
    ))
    story.append(_spacer(4))

    # Merge: analysis_partitions (Basis) + tsk_partitions (Analyse-Status) + partition_profiles (OS)
    _tsk_by_offset  = {p['offset']: p for p in (ctx.tsk_partitions or [])}
    _prof_by_offset = {p['offset']: p for p in (ctx.partition_profiles or [])}
    _ap             = ctx.analysis_partitions or []

    if _ap:
        part_rows = [['#', 'Offset (Sektoren)', 'Groesse', 'Dateisystem', 'OS / Rolle', 'Dateien', 'Status']]
        for p in sorted(_ap, key=lambda x: x.get('offset', 0)):
            off  = p.get('offset', 0)
            size = p.get('size_mb', 0)
            size_str = f'{size:.0f} MB' if size < 1024 else f'{size/1024:.2f} GB'
            tsk  = _tsk_by_offset.get(off, {})
            prof = _prof_by_offset.get(off, {})
            # OS-Rolle
            os_str = prof.get('os_name', '') or prof.get('os_family', '') or '—'
            if prof.get('is_primary'):
                os_str = f'[Primaer] {os_str}'
            # Dateien
            files_str = f"{tsk.get('files', 0):,}" if tsk else '—'
            # Status
            tsk_status = tsk.get('status', '')
            if tsk_status == 'analysiert':
                status_str = '✅ analysiert'
            elif tsk_status:
                status_str = f'⏭ {tsk_status}'
            else:
                status_str = '— nicht via TSK'
            part_rows.append([
                str(p.get('index', '?')),
                str(off),
                size_str,
                p.get('fs_type', '—'),
                os_str,
                files_str,
                status_str,
            ])
        story.append(_table(part_rows, [12*mm, 38*mm, 22*mm, 25*mm, 45*mm, 18*mm, 30*mm]))
    else:
        story.append(_body('Keine Partitionsdaten verfuegbar (Stage 02 nicht ausgefuehrt oder fehlgeschlagen).'))

    story.append(_spacer(4))

    # E01-Hashes (forensische Integritaet des Images)
    story.append(_h2('Forensische Integritaet des Beweismittels'))
    story.append(_spacer(2))
    hash_rows = [
        ['Dateiname',   ctx.disk_image_path.name if ctx.disk_image_path else '—'],
        ['SHA256',      ctx.sha256 or '—'],
        ['MD5',         ctx.md5 or '—'],
        ['Groesse',     f'{ctx.file_size_gb:.3f} GB' if ctx.file_size_gb else '—'],
        ['Format',      ctx.file_type or '—'],
    ]
    # E01-interne Hashwerte falls vorhanden
    e01_hash = getattr(ctx, 'e01_hash', '')
    e01_md5  = getattr(ctx, 'e01_md5', '')
    if e01_hash:
        hash_rows.append(['E01-interner SHA256', e01_hash])
    if e01_md5:
        hash_rows.append(['E01-interner MD5', e01_md5])
    story.append(_kv_table(hash_rows, W, _rl_color))
    story.append(PageBreak())

    # ── Seite 3: Executive Summary ─────────────────────────────────────────────
    story.append(_h1('Executive Summary'))
    story.append(_spacer(4))
    stats = [
        (f'{ctx.total_log_lines:,}',          'Log-Zeilen'),
        (f'{ctx.parsed_events:,}',             'Events'),
        (str(len(ctx.iocs)),                   'IOCs'),
        (str(len(ctx.antiforensics_hits)),     'Anti-Forensics'),
        (str(len(ctx.anomalies)),              'ML-Anomalien'),
    ]
    story.append(_stat_boxes(stats, W, _rl_color))
    story.append(_spacer(6))

    critical   = [e for e in ctx.normalized_events if e.severity == 'critical']
    high       = [e for e in ctx.normalized_events if e.severity == 'high']
    top_events = (critical + high)[:10]
    if top_events:
        story.append(_h2('Kritische Funde'))
        story.append(_spacer(2))
        rows = [['Schwere', 'Zeitstempel', 'Beschreibung', 'Quelle']]
        for e in top_events:
            rows.append([
                e.severity.upper(),
                e.timestamp.strftime('%Y-%m-%d %H:%M'),
                e.message[:85],
                e.source[:20],
            ])
        story.append(_table(rows, [25*mm, 35*mm, 90*mm, 20*mm]))
        story.append(_spacer(4))

    # Sofortmassnahmen — nur auf tatsaechlich vorhandenen Daten basierend
    # (MITRE-Empfehlungen auskommentiert: stage11_mitre nicht aktiv)
    recs = []
    if ctx.antiforensics_hits:
        recs.append('Beweise sichern — Anti-Forensik-Techniken erkannt (Details: Seite 4)')
    critical_af = [h for h in ctx.antiforensics_hits if h.get('severity') == 'critical']
    if critical_af:
        recs.append(f'{len(critical_af)} KRITISCHE Anti-Forensik-Befunde — sofortige Pruefung erforderlich')
    if ctx.tsk_fallback_used:
        recs.append('IOC-Qualitaet MITTEL — manuelle Nachpruefung empfohlen')
    if recs:
        story.append(_h2('Sofortmassnahmen'))
        story.append(_spacer(2))
        for i, r in enumerate(recs, 1):
            story.append(_body(f'{i}. {r}'))

    # ── AUSKOMMENTIERT: MITRE-Sofortmassnahmen (stage11_mitre nicht aktiv) ───
    # ctx.mitre_hits ist immer leer solange Stage 11 auskommentiert ist.
    # Zum Reaktivieren: if False: Block entfernen + stage11_mitre einkommentieren.
    if False:
        if any(h['technique_id'].startswith('T1003') for h in ctx.mitre_hits):
            recs.append('Alle Passwoerter sofort aendern (Credential Dumping erkannt)')
        if any(h['technique_id'] == 'T1098' for h in ctx.mitre_hits):
            recs.append('Neue Benutzerkonten ueberpruefen und ggf. deaktivieren')
    story.append(PageBreak())

    # ── AUSKOMMENTIERT: Alt Seite 3 — Forensische Befunde (Stage 8.5) ────────
    # stage_timeline_analysis ist in pipeline.py auskommentiert.
    # ctx.forensic_findings ist daher immer leer.
    # Reaktivieren: if False entfernen + stage_timeline_analysis einkommentieren.
    if False:
        story.append(_h1('Forensische Befunde'))
        story.append(_spacer(3))
        story.append(_body(
            'Automatisch erkannte Anomalien basierend auf MACtime-Timestamps, '
            'Anti-Forensics-Indikatoren und CVE-Zeitfenstern.'
        ))
        story.append(_spacer(4))
        if ctx.forensic_findings:
            rows = [['Schwere', 'Regel', 'Datei', 'Beschreibung']]
            for f in ctx.forensic_findings[:30]:
                rows.append([f.severity, f.rule[:25], f.file[:40], f.description[:80]])
            story.append(_table(rows, [20*mm, 35*mm, 45*mm, 70*mm]))
            story.append(_spacer(4))
            story.append(_body('Vollstaendige Befunde: forensic_findings.json'))
        else:
            story.append(_body('Keine forensischen Befunde erkannt.'))
        story.append(PageBreak())

    # ── Seite 3: Forensische Timeline ────────────────────────────────────────
    story.append(_h1('Forensische Timeline'))
    story.append(_spacer(3))
    story.append(_body(
        'Alle Zeitstempel sind in UTC. Angezeigt: kritische / hohe / mittlere Ereignisse '
        'chronologisch (max. 50). Vollstaendige Timeline: activity_timeline.csv'
    ))
    story.append(_spacer(4))
    timeline_events = sorted(
        [e for e in ctx.normalized_events if e.severity in ('critical', 'high', 'medium')],
        key=lambda e: e.timestamp
    )[:50]
    if timeline_events:
        _combined_tl = getattr(ctx, 'combined_case', False)
        _src_hdr = 'Image / Quelle' if _combined_tl else 'Quelle'
        rows = [['Zeitstempel (UTC)', 'Ereignis', _src_hdr, 'Schwere']]
        for e in timeline_events:
            _src = (_evidence_source(e)[:34] if _combined_tl else (e.source or '')[:20])
            rows.append([
                e.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                e.message[:70],
                _src,
                e.severity.upper(),
            ])
        story.append(_table(rows, [40*mm, 78*mm, 36*mm, 16*mm]))
    else:
        story.append(_body('Keine relevanten Timeline-Events.'))
    story.append(_spacer(5))
    ts_url = 'http://localhost:5000/sketch/1/explore'
    ts_box = Table([[Paragraph(
        f'<font color="#{_hex(C_MID_BLUE)}"><b>Interaktive Timeline (Timesketch)</b></font>'
        f'<br/><font size="9">{ts_url}</font>',
        _normal_style()
    )]], colWidths=[W])
    ts_box.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), _rl_color_from_tuple(C_LIGHT_BLUE)),
        ('BOX',           (0, 0), (-1, -1), 1.5, _rl_color_from_tuple(C_MID_BLUE)),
        ('TOPPADDING',    (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING',   (0, 0), (-1, -1), 10),
    ]))
    story.append(ts_box)
    story.append(PageBreak())

    # ── Seite 4: Anti-Forensics Befundbericht (Betreuer-Template-Format) ─────
    story.append(_h1('Anti-Forensics — Befundbericht'))
    story.append(_spacer(3))
    story.append(_body(
        'Die folgenden Befunde basieren auf forensischer Analyse des Disk-Images via '
        'TSK/icat ohne Mounting (forensisch unveraendert). Jeder Abschnitt enthaelt '
        'eine technische Einordnung (Zusammenspiel) und den konkreten Befund.'
    ))
    story.append(_spacer(5))

    # Hilfsfunktionen fuer Befund-Boxen im Betreuer-Template-Format
    def _befund_pos(text):
        """Rote Box: Positiv-Befund (verdaechtig / Anti-Forensik-Indikator gefunden)."""
        b = Table([[Paragraph(
            f'<font size="8" color="#{_hex(C_RED)}"><b>Befund (Positiv):</b></font> '
            f'<font size="8" color="#{_hex(C_DARK_GREY)}">{text}</font>',
            _normal_style_left()
        )]], colWidths=[W])
        b.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, -1), _rl_color_from_tuple(C_RED_L)),
            ('BOX',           (0, 0), (-1, -1), 1.0, _rl_color_from_tuple(C_RED)),
            ('LEFTPADDING',   (0, 0), (-1, -1), 8),
            ('TOPPADDING',    (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        return b

    def _befund_neg(text):
        """Gruene Box: Negativ-Befund (sauber / kein Indikator gefunden)."""
        b = Table([[Paragraph(
            f'<font size="8" color="#{_hex(C_GREEN)}"><b>Befund (Negativ):</b></font> '
            f'<font size="8" color="#{_hex(C_DARK_GREY)}">{text}</font>',
            _normal_style_left()
        )]], colWidths=[W])
        b.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, -1), _rl_color_from_tuple(C_GREEN_L)),
            ('BOX',           (0, 0), (-1, -1), 1.0, _rl_color_from_tuple(C_GREEN)),
            ('LEFTPADDING',   (0, 0), (-1, -1), 8),
            ('TOPPADDING',    (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ]))
        return b

    def _ref(text):
        return _body(f'<i>Ref.: {text}</i>')

    def _zsp(text):
        return _body(f'<b>Zusammenspiel:</b> {text}')

    grub    = getattr(ctx, 'grub_config', {})
    symlinks = getattr(ctx, 'primary_symlinks', {})

    # ── Log-Manipulation ─────────────────────────────────────────────────────
    story.append(_h2('Log-Manipulation'))
    story.append(_spacer(2))

    devnull_hits  = [h for h in ctx.antiforensics_hits if h.get('type') == 'devnull_symlink']
    devnull_paths = [p for p, t in symlinks.items() if '/dev/null' in t]
    if devnull_hits:
        paths_str = ', '.join(devnull_paths[:5])
        story.append(_befund_pos(
            f'Die folgenden Log-Dateien bzw. Shell-Histories sind via Symlink auf /dev/null '
            f'gesetzt: {paths_str}. Somit werden keinerlei Protokolle geschrieben.'
        ))
    else:
        story.append(_befund_neg(
            'Es wurden keine relevanten Log-Dateien oder Shell-Histories via Symlink auf '
            '/dev/null gesetzt. Die Protokollierung ist nicht auf diesem Wege unterbunden.'
        ))
    story.append(_spacer(2))

    wtmp_hits = [h for h in ctx.antiforensics_hits if h.get('type') == 'journal_wtmp_mismatch']
    if wtmp_hits:
        story.append(_befund_pos(wtmp_hits[0].get('details', '')[:300]))
    else:
        story.append(_befund_neg(
            'Die Login-Ereignisse aus wtmp und dem systemd-Journal stimmen ueberein. '
            'Es liegen keine Anzeichen fuer nachtraegliche Log-Manipulation vor.'
        ))
    story.append(_spacer(3))

    # ── rc.local ─────────────────────────────────────────────────────────────
    story.append(_h2('Automatische Aktionen beim Systemstart (rc.local / local.d)'))
    story.append(_spacer(2))
    rc_hits = [h for h in ctx.antiforensics_hits if h.get('type') == 'rc_local_antiforensics']
    if rc_hits:
        details_str = '; '.join(h.get('details', '')[:80] for h in rc_hits[:3])
        story.append(_befund_pos(
            f'In der Startkonfiguration wurden Anti-Forensik-Kommandos identifiziert: '
            f'{details_str}.'
        ))
    else:
        story.append(_befund_neg(
            'Es wurden keine Anti-Forensik-Kommandos (Log-Loeschung, /dev/null-Verlinkung, '
            'Service-Stopp) in der Systemstartdatei festgestellt.'
        ))
    story.append(_spacer(2))
    story.append(_ref('/etc/rc.local, /etc/rc.d/rc.local, /etc/local.d/*.start'))
    story.append(_spacer(4))

    # ── 1. GRUB Bootloader Konfiguration ─────────────────────────────────────
    story.append(_h2('1. GRUB Bootloader Konfiguration (Boot-Parameter)'))
    story.append(_spacer(2))
    story.append(_zsp(
        'Der Bootloader uebergibt beim Systemstart Parameter an den Kernel, welche das '
        'Speichermanagement priorisiert steuern (z.B. sofortiges Nullen freigegebener Pages '
        'via init_on_free), noch bevor einkompilierte Standardwerte greifen.'
    ))
    story.append(_spacer(2))
    grub_af     = grub.get('antiforensic_params', [])
    grub_hits   = [h for h in ctx.antiforensics_hits if h.get('type') == 'grub_memory_wipe']
    if grub_hits:
        params_str = ', '.join(grub_af) if grub_af else 'init_on_free / page_poison'
        story.append(_befund_pos(
            f'Im Rahmen der Auswertung wurden Boot-Parameter festgestellt, die eine '
            f'sofortige Bereinigung freigegebener Speicherseiten erzwingen '
            f'(Parameter: {params_str}). Dies erschwert die post-mortem Rekonstruktion '
            f'volatiler Daten aus Speicherdumps erheblich.'
        ))
    else:
        story.append(_befund_neg(
            'Es wurden keine dynamischen Boot-Parameter zur aktiven Bereinigung des '
            'Arbeitsspeichers (z.B. init_on_free, page_poison) festgestellt.'
        ))
    sources = grub.get('sources', [])
    story.append(_spacer(2))
    story.append(_ref(', '.join(sources) if sources
                     else '/etc/default/grub, /boot/grub/grub.cfg, /boot/grub/grubenv'))
    story.append(_spacer(4))

    # ── 2. Hardcoded Kernel Konfigurationen ───────────────────────────────────
    story.append(_h2('2. Hardcoded Kernel Konfigurationen'))
    story.append(_spacer(2))
    story.append(_zsp(
        'Die statische Konfigurationsdatei spiegelt die Kompilierungs-Flags des Kernels '
        'wider. Fehlen dynamische GRUB-Parameter, definiert diese Basis '
        '(z.B. CONFIG_INIT_ON_FREE_DEFAULT_ON=y), ob das OS Speicherseiten nativ '
        'beim Freigeben ueberschreibt.'
    ))
    story.append(_spacer(2))
    flags_map   = getattr(ctx, 'kernel_compile_flags', {})
    flags_hits  = [h for h in ctx.antiforensics_hits if h.get('type') == 'kernel_compile_antiforensics']
    all_flags   = list(dict.fromkeys(
        f for info in flags_map.values() for f in info.get('active_flags', [])
    ))
    if flags_hits and all_flags:
        story.append(_befund_pos(
            f'Die Analyse der statischen Kernel-Konfiguration weist einkompilierte '
            f'Anti-Forensik-Flags auf: {", ".join(all_flags)}. '
            f'Der Kernel ueberschreibt Speicherbereiche nativ bei deren Freigabe.'
        ))
    else:
        story.append(_befund_neg(
            'Die einkompilierte Kernel-Konfiguration weist keine aktiven Memory-Wiping-Flags '
            'auf. Eine systemseitige, automatische Ueberschreibung des RAMs ist auf '
            'Kernel-Ebene nicht konfiguriert.'
        ))
    all_k = getattr(ctx, 'all_kernel_versions', []) or [ctx.kernel_version]
    story.append(_spacer(2))
    story.append(_ref(', '.join(f'/boot/config-{k}' for k in all_k[:3])))
    story.append(_spacer(4))

    # ── 3. Systemd Services (Userland Anti-Forensik & Persistenz) ───────────
    story.append(_h2('3. Systemd Services (Userland Anti-Forensik & Persistenz)'))
    story.append(_spacer(2))
    story.append(_zsp(
        'Systemd steuert den Lebenszyklus von Diensten. Ueber die ExecStop-Direktive '
        'koennen bei Terminierung eines Prozesses gezielt externe Userland-Tools '
        '(z.B. sdmem) aufgerufen werden, um den Applikationsspeicher zu wipen. '
        'Darueber hinaus sind Dienste ausserhalb des OS-Standard-Sets ein Indikator '
        'fuer persistente Software, Backdoors oder Angreifer-Infrastruktur.'
    ))
    story.append(_spacer(2))

    # 3a: ExecStop-Wiping
    execstop_hits = [h for h in ctx.antiforensics_hits if h.get('type') == 'execstop_wiping']
    if execstop_hits:
        story.append(_befund_pos(
            f'In der Dienstkonfiguration wurde eine gezielte Anti-Forensik-Routine '
            f'identifiziert. Beim Beenden des Dienstes wird ein Wiping-Tool via ExecStop '
            f'ausgefuehrt: {execstop_hits[0].get("details", "")[:180]}'
        ))
    else:
        story.append(_befund_neg(
            'Die Ueberpruefung der relevanten Service-Units ergab keine Aufrufe von '
            'externen Speicher-Wiping-Tools bei Prozessterminierungen. '
            'Verdaechtige ExecStop-Direktiven liegen nicht vor.'
        ))
    story.append(_spacer(2))

    # 3b: Non-Standard Services
    _primary_pf_af = next((p for p in ctx.partition_profiles if p.get('is_primary')), {})
    _svc_af        = _primary_pf_af.get('services', {})
    _non_std_af    = _svc_af.get('non_standard', [])
    if _non_std_af:
        ns_str_af = ', '.join(_non_std_af[:12])
        if len(_non_std_af) > 12:
            ns_str_af += f' (+{len(_non_std_af) - 12} weitere)'
        story.append(_befund_pos(
            f'{len(_non_std_af)} Nicht-Standard-Service(s) wurden aktiviert vorgefunden, '
            f'die nicht zum erwarteten Standard-Set der OS-Familie gehoeren: {ns_str_af}. '
            f'Diese Dienste sind manuell auf Legitimitaet und potenzielle Persistenz-Mechanismen '
            f'zu pruefen.'
        ))
    else:
        story.append(_befund_neg(
            'Alle aktivierten systemd-Services entsprechen dem erwarteten Standard-Set der '
            'erkannten OS-Familie. Es wurden keine nicht-standardmaessigen Dienste vorgefunden.'
        ))
    story.append(_spacer(2))
    story.append(_ref(
        '/etc/systemd/system/*.service, /lib/systemd/system/*.service, '
        '/usr/lib/systemd/system/*.service'
    ))
    story.append(_spacer(4))

    # ── 4. Swap Konfiguration ─────────────────────────────────────────────────
    story.append(_h2('4. Swap Konfiguration'))
    story.append(_spacer(2))
    story.append(_zsp(
        'Bei hohem RAM-Bedarf lagert der Kernel inaktive Speicherseiten auf einen '
        'persistenten Datentraeger aus (Swap). Diese Fragmente sind essenzielle Quellen '
        'fuer die Deadbox-Speicherforensik.'
    ))
    story.append(_spacer(2))
    swap = getattr(ctx, 'swap_config', {})
    if swap.get('found'):
        entries   = swap.get('entries', [])
        swap_info = ', '.join(f"{e.get('type','?')}: {e.get('path','?')}" for e in entries[:3])
        story.append(_befund_pos(
            f'Es wurde eine aktive Auslagerungskonfiguration ({swap_info}) festgestellt. '
            f'Diese stellt eine primaere Quelle fuer ausgelagerte volatile Artefakte dar.'
        ))
    else:
        story.append(_befund_neg(
            'Es ist weder eine Swap-Partition noch eine Swap-Datei konfiguriert. '
            'Das System lagert keine Speicherseiten auf den persistenten Datentraeger aus. '
            'Eine Deadbox-Speicherforensik ueber Swap-Artefakte ist nicht moeglich.'
        ))
    story.append(_spacer(2))
    story.append(_ref('/etc/fstab'))
    story.append(_spacer(4))

    # ── 5. Systemzustand (Kernel-Diskrepanz & Pending Reboot) ─────────────────
    story.append(_h2('5. Systemzustand (Kernel-Diskrepanz & Pending Reboot)'))
    story.append(_spacer(2))
    story.append(_zsp(
        'Paketmanager schreiben nach Kernel-Updates die grub.cfg neu. Der geladene Kernel '
        'aendert sich jedoch erst nach einem Reboot. Diskrepanzen zwischen Konfiguration '
        'und System-Logs belegen die Uptime auf einem Alt-Kernel.'
    ))
    story.append(_spacer(2))
    disc_hits      = [h for h in ctx.antiforensics_hits if h.get('type') == 'kernel_discrepancy']
    reboot_hits    = [h for h in ctx.antiforensics_hits if h.get('type') == 'pending_reboot']
    active_k       = grub.get('active_kernel', '')
    loaded_k       = getattr(ctx, 'loaded_kernel_from_logs', '')
    reboot_pending = getattr(ctx, 'reboot_pending', False)
    if disc_hits:
        story.append(_befund_pos(
            f'Es liegt eine Diskrepanz zwischen dem konfigurierten Default-Kernel '
            f'({active_k or "unbekannt"}) und dem tatsaechlich geladenen Kernel '
            f'({loaded_k or "unbekannt"}) vor.'
            + (' Das Vorhandensein des Reboot-Flags belegt, dass das System seit dem '
               'Kernel-Update nicht neu gestartet wurde.'
               if reboot_pending else '')
        ))
    elif reboot_hits:
        story.append(_befund_pos(
            'Das System hat ein ausstehendes Kernel-Update — das Reboot-Flag ist gesetzt, '
            'aber das System wurde noch nicht neu gestartet. '
            f'Aktuell geladen laut Logs: {loaded_k or active_k or "unbekannt"}. '
            'Es liegt keine Diskrepanz zwischen GRUB-Konfiguration und geladener Version vor.'
        ))
    else:
        story.append(_befund_neg(
            f'Der in der GRUB-Konfiguration hinterlegte Default-Kernel '
            f'({active_k or ctx.kernel_version or "unbekannt"}) stimmt mit dem '
            f'ermittelten Systemstatus ueberein. Es liegen keine Artefakte vor, die auf '
            f'einen ausstehenden Neustart oder eine Laufzeitdiskrepanz hindeuten.'
        ))
    story.append(_spacer(2))
    story.append(_ref(
        '/boot/grub/grub.cfg, /var/run/reboot-required, '
        '/var/log/kern.log, /var/log/messages, /var/log/journal/*.journal '
        '(via journalctl --file, inkl. LZ4-komprimierte Journals)'
    ))
    story.append(PageBreak())

    # ── AUSKOMMENTIERT: Alt Seite 5 — Anti-Forensics Tabelle + ML ────────────
    # Ersetzt durch Seite 4 im Betreuer-Template-Format.
    # ML-Anomalien: stage10_ml auskommentiert — ctx.anomalies immer leer.
    # Reaktivieren: if False entfernen.
    if False:
        story.append(_h1('Anti-Forensics & ML-Anomalien'))
        story.append(_spacer(3))
        if ctx.antiforensics_hits:
            story.append(_body(f'Anti-Forensics: {len(ctx.antiforensics_hits)} Treffer'))
            rows = [['Technik', 'Datei / Quelle', 'Details', 'Schwere']]
            for h in ctx.antiforensics_hits[:20]:
                rows.append([h['type'], h['file'][:40], h['details'][:60], h['severity']])
            story.append(_table(rows, [35*mm, 45*mm, 70*mm, 20*mm]))
        story.append(_h2('ML-Anomalien (Isolation Forest)'))
        if ctx.anomalies:
            rows = [['Score', 'Zeitstempel', 'Event', 'Quelle']]
            for e in sorted(ctx.anomalies, key=lambda x: x.anomaly_score, reverse=True)[:20]:
                rows.append([f'{e.anomaly_score:.2f}',
                             e.timestamp.strftime('%Y-%m-%d %H:%M'),
                             e.message[:70], e.source[:20]])
            story.append(_table(rows, [18*mm, 35*mm, 95*mm, 22*mm]))
        else:
            story.append(_body('Keine Anomalien erkannt.'))
        story.append(PageBreak())

    # ── Seite 5: IOC-Liste ───────────────────────────────────────────────────
    story.append(_h1('Indicators of Compromise (IOCs)'))
    story.append(_spacer(3))
    story.append(_body(
        f'IOC-Qualitaet: {ctx.ioc_quality}  |  {len(ctx.iocs)} IOCs extrahiert  |  '
        f'Vollstaendige Liste: iocs.json'
    ))
    story.append(_spacer(4))

    ips     = [i for i in ctx.iocs if i.type in ('ip', 'ipv6')]
    domains = [i for i in ctx.iocs if i.type in ('domain', 'url')]
    hashes  = [i for i in ctx.iocs if i.type in ('hash_md5', 'hash_sha256')]
    others  = [i for i in ctx.iocs if i.type in ('email', 'cve', 'registry_key')]

    for group_name, group in [('IP-Adressen', ips), ('Domains / URLs', domains),
                               ('Datei-Hashes', hashes), ('Sonstige IOCs', others)]:
        if not group:
            continue
        story.append(_h2(group_name))
        story.append(_spacer(2))
        rows = [['Typ', 'Wert', 'Quelle']]
        for ioc in group[:30]:
            rows.append([ioc.type, ioc.value[:60], ioc.source[:20]])
        story.append(_table(rows, [25*mm, 117*mm, 28*mm]))
        story.append(_spacer(4))
    if not (ips or domains or hashes or others):
        story.append(_body('Keine IOCs extrahiert.'))

    story.append(PageBreak())

    # ── Seite 6: Nutzer-Profil & Netzwerk ────────────────────────────────────
    story.append(_h1('Nutzer-Profil & Netzwerk'))
    story.append(_spacer(3))
    primary_profile = next((p for p in ctx.partition_profiles if p.get('is_primary')), {})

    # Nutzer-Tabelle
    story.append(_h2('Nutzer-Uebersicht'))
    story.append(_spacer(2))
    users = ctx.users or primary_profile.get('users', [])
    display_users = [u for u in users if not u.get('is_system') or u.get('is_unexpected')]
    if display_users:
        rows = [['Name', 'UID', 'Shell', 'Login', 'Passwort', 'Sudo', 'Auffaellig']]
        for u in display_users[:15]:
            flag = ''
            if u.get('uid') == 0 and u.get('name') != 'root':
                flag = '! UID=0'
            elif u.get('is_unexpected'):
                flag = '! unbekannt'
            rows.append([
                u.get('name', ''),
                str(u.get('uid', '')),
                (u.get('shell', '') or '')[-20:],
                'Ja' if u.get('login_allowed') else 'Nein',
                'Ja' if u.get('has_password')  else 'Nein',
                'Ja' if u.get('has_sudo')       else '-',
                flag,
            ])
        story.append(_table(rows, [28*mm, 14*mm, 35*mm, 14*mm, 18*mm, 14*mm, 30*mm]))
    else:
        story.append(_body('Keine regulaeren Nutzer gefunden.'))
    story.append(_spacer(4))

    # SSH-Konfiguration
    ssh = primary_profile.get('ssh_config', {})
    if ssh:
        story.append(_h2('SSH-Konfiguration'))
        story.append(_spacer(2))
        ssh_rows = []
        if ssh.get('permit_root_login'):
            ssh_rows.append(['Root-Login', ssh['permit_root_login']])
        if ssh.get('password_auth'):
            ssh_rows.append(['Passwort-Authentifizierung', ssh['password_auth']])
        if ssh.get('port', '') not in ('22', ''):
            ssh_rows.append(['SSH-Port (nicht standard)', ssh['port']])
        if ssh.get('allow_users'):
            ssh_rows.append(['AllowUsers', ssh['allow_users']])
        if ssh.get('deny_users'):
            ssh_rows.append(['DenyUsers', ssh['deny_users']])
        if ssh_rows:
            story.append(_kv_table(ssh_rows, W, _rl_color))
        story.append(_spacer(4))

    # Netzwerk
    net = primary_profile.get('network', {})
    story.append(_h2('Netzwerk-Konfiguration'))
    story.append(_spacer(2))
    net_rows = []
    if ctx.ip_addresses:
        net_rows.append(['IP-Adressen', ', '.join(ctx.ip_addresses[:8])])
    if net.get('interfaces'):
        net_rows.append(['Interfaces', ', '.join(net['interfaces'][:6])])
    if net.get('gateway'):
        net_rows.append(['Gateway', net['gateway']])
    if net.get('dns_servers'):
        net_rows.append(['DNS-Server', ', '.join(net['dns_servers'][:4])])
    if net.get('mac_hints'):
        net_rows.append(['MAC-Hinweise', ', '.join(net['mac_hints'][:3]) + ' (aus DHCP/NM)'])
    if net_rows:
        story.append(_kv_table(net_rows, W, _rl_color))
    else:
        story.append(_body('Netzwerk-Daten nicht verfuegbar.'))
    story.append(PageBreak())

    # ── Seite 7: System-Profiling ─────────────────────────────────────────────
    story.append(_h1('System-Profiling'))
    story.append(_spacer(4))

    # ── Fall-Modus: ein Profil-Block je Image ─────────────────────────────
    _evi_sp = getattr(ctx, 'evidence_items', []) or []
    if getattr(ctx, 'combined_case', False) and _evi_sp:
        for ev in _evi_sp:
            story.append(_h2(f"Image: {ev.get('name', '-')}"))
            story.append(_spacer(2))
            pps = ev.get('partition_profiles') or [{}]
            pp  = next((p for p in pps if p.get('is_primary')), pps[0] if pps else {})
            _lbl = f"Partition {pp.get('partition_index','?')} (offset {pp.get('offset','?')})"
            rows = [
                ['OS-Name',     ev.get('os_name') or '-'],
                ['OS-Familie',  ev.get('os_family') or '-'],
                ['OS-Quelle',   pp.get('os_source', '') or 'target-query'],
                ['Kernel',      ev.get('kernel_version') or '-'],
                ['Hostname',    ev.get('hostname') or '-'],
                ['Zeitzone',    ev.get('timezone_display') or ev.get('timezone') or '-'],
                ['Machine-ID',  (f"{ev.get('machine_id')}   [/etc/machine-id, {_lbl}]"
                                 if ev.get('machine_id') else '-')],
                ['/etc/shadow geaendert',
                 (f"{ev.get('shadow_mtime')}   [TSK istat, {_lbl}] — letzte Passwort-/Benutzeraenderung"
                  if ev.get('shadow_mtime') else '-')],
                ['Partitionen', str(len(ev.get('partition_layout') or []))],
            ]
            story.append(_kv_table(rows, W, _rl_color))
            story.append(_spacer(5))
        story.append(PageBreak())
    else:
        loaded_k_s7      = getattr(ctx, 'loaded_kernel_from_logs', '') or '-'
        all_kernels_s7   = getattr(ctx, 'all_kernel_versions', [])
        # Provenienz-Label der primaeren Partition (Review-Punkt #7:
        # jeder Wert mit nachpruefbarer Quelle — Datei, Methode, Partition)
        _pp_idx = primary_profile.get('partition_index', '?')
        _pp_off = primary_profile.get('offset', '?')
        _pp_lbl = f'Partition {_pp_idx} (offset {_pp_off})'
        _os_src = primary_profile.get('os_source', '')
        kv3 = [
            ['OS-Familie',               ctx.os_family or '-'],
            ['OS-Name',                  ctx.os_name or '-'],
            ['OS-Erkennungsquelle',      _os_src or 'target-query (ganzes Image)'],
            ['Kernel (target-query)',    ctx.kernel_version or '-'],
            ['Kernel (geladen, Logs)',   loaded_k_s7],
            ['Alle installierten Kernel', ', '.join(all_kernels_s7) or '-'],
            ['Hostname',                 ctx.hostname or '-'],
            ['Zeitzone',                 getattr(ctx, 'timezone_display', None) or ctx.timezone or '-'],
            ['Machine-ID',               (f'{ctx.machine_id}   [/etc/machine-id via TSK icat, {_pp_lbl}]'
                                          if ctx.machine_id else '-')],
            ['/etc/shadow geaendert',    (f'{ctx.shadow_mtime}   [TSK istat auf /etc/shadow, {_pp_lbl}] '
                                          f'— letzte Passwort-/Benutzeraenderung'
                                          if ctx.shadow_mtime else '-')],
            ['Image-Format',             ctx.file_type],
            ['Image-Groesse',            f'{ctx.file_size_gb:.2f} GB'],
        ]
        if getattr(ctx, 'sha1', ''):
            kv3.append(['SHA1 (E01-eingebettet)', ctx.sha1])
        if ctx.sha256:
            kv3.append(['SHA256', ctx.sha256])
        kv3 += [
            ['MD5',                      ctx.md5 or '-'],
            ['Virtualisierung',          primary_profile.get('virtualization', '-')],
        ]
        story.append(_kv_table(kv3, W, _rl_color))
        story.append(_spacer(4))

        # Non-Standard Services — forensisch relevante Persistenz-Hinweise
        _svc_s7       = primary_profile.get('services', {})
        _non_std_s7   = _svc_s7.get('non_standard', [])
        _enabled_s7   = _svc_s7.get('enabled', [])
        story.append(_h2('Aktivierte Services'))
        story.append(_spacer(2))
        if _enabled_s7:
            story.append(_body(
                f'Aktivierte systemd-Services ({len(_enabled_s7)} gesamt): '
                f'{", ".join(_enabled_s7[:12])}'
                + (f' (+{len(_enabled_s7) - 12} weitere)' if len(_enabled_s7) > 12 else '')
            ))
        else:
            story.append(_body('Keine aktivierten Services erkannt (ggf. OpenRC-System).'))
        story.append(_spacer(2))
        if _non_std_s7:
            ns_str = ', '.join(_non_std_s7[:15])
            if len(_non_std_s7) > 15:
                ns_str += f' (+{len(_non_std_s7) - 15} weitere)'
            story.append(_body(
                f'⚠ {len(_non_std_s7)} Nicht-Standard-Service(s) erkannt '
                f'(nicht in OS-Whitelist): {ns_str}. '
                f'Diese Dienste koennen auf Persistenz, Backdoors oder installierte '
                f'Drittanbieter-Software hinweisen und sind forensisch zu pruefen.'
            ))
        else:
            story.append(_body(
                'Alle aktivierten Services entsprechen dem erwarteten Standard-Set '
                'der erkannten OS-Familie. Es wurden keine abweichenden Dienste festgestellt.'
            ))
        story.append(PageBreak())

        # ── Seite 8: Basic Checks — Log-Praesenz & Konsistenz ────────────────────
    story.append(_h1('Basic Checks — Log-Praesenz & Konsistenz'))
    story.append(_spacer(3))
    story.append(_body(
        'Automatische Ueberpruefung: (1) Pflicht-Logs pro OS-Familie vorhanden? '
        '(2) Installierte Dienste haben erwartete Log-Dateien? '
        '(3) Log-Dateien ohne zugehoerige Installation vorhanden (Anomalie)?'
    ))
    story.append(_spacer(4))

    _bc       = getattr(ctx, 'basic_checks', [])
    _bc_anom  = getattr(ctx, 'basic_check_anomalies', 0)

    if not _bc:
        story.append(_body(
            f'Basic Checks nicht ausgefuehrt — OS-Familie "{ctx.os_family}" '
            f'hat kein Pruef-Profil oder Stage 03.5 wurde uebersprungen.'
        ))
    else:
        # Zusammenfassung-Box
        _bc_total = len(_bc)
        _bc_ok    = sum(1 for c in _bc if not c.get('anomaly') and c.get('found'))
        if _bc_anom == 0:
            bc_summary_bg     = C_GREEN_L
            bc_summary_border = C_GREEN
            bc_summary_text   = f'Alle {_bc_total} Checks ohne Anomalie — keine verdaechtigen Log-Abweichungen erkannt.'
        else:
            bc_summary_bg     = C_ORANGE_L
            bc_summary_border = C_ORANGE
            bc_summary_text   = (
                f'{_bc_anom} Anomalie(n) in {_bc_total} geprueften Eintraegen erkannt. '
                f'Details: rote / gelbe Zeilen in der Tabelle unten.'
            )
        bc_box = Table([[Paragraph(
            f'<font size="9"><b>{bc_summary_text}</b></font>',
            _normal_style_left()
        )]], colWidths=[W])
        bc_box.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, -1), _rl_color_from_tuple(bc_summary_bg)),
            ('BOX',           (0, 0), (-1, -1), 1.5, _rl_color_from_tuple(bc_summary_border)),
            ('TOPPADDING',    (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING',   (0, 0), (-1, -1), 10),
        ]))
        story.append(bc_box)
        story.append(_spacer(4))

        # Tabelle: alle Checks — manuell gebaut damit Anomalie-Farben
        # in einem setStyle()-Aufruf gesetzt werden koennen (kein ._style-Zugriff).
        from reportlab.platypus import Table as _BCT, TableStyle as _BCTS
        from reportlab.lib.styles import ParagraphStyle as _BCPS
        _bc_col_w = [35*mm, 52*mm, 18*mm, 18*mm, 47*mm]
        _bc_hs = _BCPS('bc_h', fontSize=8, fontName='Helvetica-Bold',
                        textColor=_rl_color_from_tuple(C_WHITE), leading=11)
        _bc_cs = _BCPS('bc_c', fontSize=8, fontName='Helvetica',
                        textColor=_rl_color_from_tuple(C_DARK_GREY), leading=11)

        bc_header = ['Service / Log', 'Pfad', 'Erwartet', 'Gefunden', 'Anomalie']
        bc_data   = [[Paragraph(str(h), _bc_hs) for h in bc_header]]
        for c in _bc:
            erw_txt  = 'Pflicht' if c.get('expected') else 'Bedingt'
            gef_txt  = '✅ Ja'   if c.get('found')    else '❌ Nein'
            anom_txt = c.get('anomaly', '') or '—'
            bc_data.append([
                Paragraph(c.get('service',  '—'),      _bc_cs),
                Paragraph(c.get('log_path', '—'),      _bc_cs),
                Paragraph(erw_txt,                      _bc_cs),
                Paragraph(gef_txt,                      _bc_cs),
                Paragraph(anom_txt[:80],                _bc_cs),
            ])

        # Alle Style-Commands auf einmal — Basis + Anomalie-Farben
        bc_style_cmds = [
            ('BACKGROUND',    (0, 0), (-1, 0),  _rl_color_from_tuple(C_DARK_BLUE)),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1),
             [_rl_color_from_tuple(C_WHITE), _rl_color_from_tuple(C_LIGHT_GREY)]),
            ('GRID',          (0, 0), (-1, -1), 0.3, _rl_color_from_tuple((0xDD/255,)*3)),
            ('LEFTPADDING',   (0, 0), (-1, -1), 6),
            ('RIGHTPADDING',  (0, 0), (-1, -1), 6),
            ('TOPPADDING',    (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('VALIGN',        (0, 0), (-1, -1), 'TOP'),
        ]
        for i, c in enumerate(_bc, 1):   # i=1: erste Datenzeile (0 = Header)
            atype = c.get('anomaly_type', '')
            if atype == 'mandatory_missing':
                bc_style_cmds += [
                    ('BACKGROUND', (0, i), (-1, i), _rl_color_from_tuple(C_RED_L)),
                    ('TEXTCOLOR',  (0, i), (-1, i), _rl_color_from_tuple(C_RED)),
                ]
            elif atype in ('install_without_log', 'log_without_install'):
                bc_style_cmds += [
                    ('BACKGROUND', (0, i), (-1, i), _rl_color_from_tuple(C_ORANGE_L)),
                    ('TEXTCOLOR',  (0, i), (-1, i), _rl_color_from_tuple(C_ORANGE)),
                ]
        bc_tbl = _BCT(bc_data, colWidths=_bc_col_w, repeatRows=1)
        bc_tbl.setStyle(_BCTS(bc_style_cmds))
        story.append(bc_tbl)

        # Legende
        story.append(_spacer(3))
        story.append(_body(
            '<font color="#{}">■</font> Pflicht-Log fehlt (kritisch)  '
            '<font color="#{}">■</font> Dienst installiert ohne Log / Log ohne Dienst (verdaechtig)  '
            '<font color="#{}">■</font> Unauffaellig'.format(
                _hex(C_RED), _hex(C_ORANGE), _hex(C_GREEN)
            )
        ))

    story.append(PageBreak())

    # ══ ANHANG ════════════════════════════════════════════════════════════════

    # ── Anhang A: Pipeline-Status & Qualitaet ─────────────────────────────────
    story.append(_banner('ANHANG', C_MID_GREY, C_WHITE, 14, W))
    story.append(_spacer(4))
    story.append(_h1('A — Pipeline-Status & Qualitaet'))
    story.append(_spacer(3))
    story.append(_body(f'Gesamtqualitaet: {quality}  |  Fehler: {len(ctx.stage_errors)}'))
    story.append(_spacer(3))
    stage_rows = [['Stufe', 'Status', 'Fehler / Anmerkung']]
    for stage_name, status in ctx.stage_status.items():
        err = ctx.stage_errors.get(stage_name, '')
        stage_rows.append([stage_name, status, err[:60]])
    story.append(_table(stage_rows, [35*mm, 40*mm, 95*mm]))
    story.append(PageBreak())

    # ── AUSKOMMENTIERT: Alt Seite 6 — Anti-Forensics Systemkonfiguration ─────
    # Tabellenbasierter Ansatz — ersetzt durch Seite 4 (Betreuer-Template-Format).
    # Reaktivieren: if False entfernen.
    if False:  # ── Alt Seite 6: Anti-Forensics Systemkonfiguration (tabellenbasiert) ──────────
        # Reaktivieren: if False entfernen
        story.append(_h1('Anti-Forensics — Systemkonfiguration'))
        story.append(_spacer(3))
        story.append(_body(
            'Systemkonfigurationsdaten die auf Anti-Forensik-Massnahmen hinweisen. '
            'Extrahiert via TSK/icat aus dem Disk-Image ohne Mounting (forensisch unveraendert).'
        ))
        story.append(_spacer(5))

        # ── GRUB-Konfiguration ────────────────────────────────────────────────────
        story.append(_h2('GRUB-Bootloader-Konfiguration'))
        story.append(_spacer(2))
        grub = getattr(ctx, 'grub_config', {})
        grub_rows = []
        if grub.get('active_kernel'):
            grub_rows.append(['Aktiver Kernel (GRUB-Default)', grub['active_kernel']])
        if grub.get('fallback_kernels'):
            grub_rows.append(['Fallback-Kernel', ', '.join(grub['fallback_kernels'])])
        if grub.get('grubenv_entry'):
            grub_rows.append(['grubenv saved_entry', grub['grubenv_entry']])
        if grub.get('boot_params'):
            grub_rows.append(['GRUB_CMDLINE_LINUX_DEFAULT', grub['boot_params'][:120]])
        if grub.get('antiforensic_params'):
            grub_rows.append(['⚠ Anti-Forensik-Parameter', ', '.join(grub['antiforensic_params'])])
        else:
            grub_rows.append(['Anti-Forensik-Parameter', 'Keine erkannt'])
        if grub.get('sources'):
            grub_rows.append(['Quellen', ', '.join(grub['sources'])])
        if grub_rows:
            from reportlab.platypus import Table as RLTable, TableStyle as RLTS
            from reportlab.lib.colors import HexColor
            data = [[r[0], r[1]] for r in grub_rows]
            tbl = RLTable(data, colWidths=[60*mm, W - 60*mm])
            style_cmds = [
                ('BACKGROUND',   (0, 0), (0, -1), _rl_color_from_tuple(C_LIGHT_GREY)),
                ('FONTNAME',     (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME',     (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE',     (0, 0), (-1, -1), 8),
                ('GRID',         (0, 0), (-1, -1), 0.3, _rl_color_from_tuple((0xDD/255,)*3)),
                ('LEFTPADDING',  (0, 0), (-1, -1), 5),
                ('TOPPADDING',   (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING',(0, 0), (-1, -1), 4),
            ]
            # Anti-Forensik-Zeile rot markieren
            for idx, r in enumerate(grub_rows):
                if '⚠' in r[0]:
                    style_cmds.append(('BACKGROUND', (0, idx), (-1, idx),
                                       _rl_color_from_tuple(C_RED_L)))
                    style_cmds.append(('TEXTCOLOR',  (0, idx), (-1, idx),
                                       _rl_color_from_tuple(C_RED)))
            tbl.setStyle(RLTS(style_cmds))
            story.append(tbl)
        else:
            story.append(_body('Keine GRUB-Konfiguration gefunden (kein grub.cfg / grubenv im Image).'))
        story.append(_spacer(5))

        # ── Kernel Compile-Flags ──────────────────────────────────────────────────
        story.append(_h2('Einkompilierte Kernel-Flags (/boot/config-*)'))
        story.append(_spacer(2))
        flags_map = getattr(ctx, 'kernel_compile_flags', {})
        all_kernels = getattr(ctx, 'all_kernel_versions', [])
        if flags_map:
            rows = [['Kernel-Version', 'Aktive Anti-Forensik-Flags', 'Bewertung']]
            for kver in all_kernels:
                info = flags_map.get(kver)
                if info is None:
                    rows.append([kver, '—  (keine /boot/config gefunden)', 'Unbekannt'])
                elif info.get('has_antiforensics'):
                    rows.append([kver,
                                 '\n'.join(info['active_flags']),
                                 '⚠ VERDAECHTIG'])
                else:
                    rows.append([kver, 'Keine Anti-Forensik-Flags aktiv', 'Unauffaellig'])
            tbl = _table(rows, [55*mm, 95*mm, 20*mm])
            # Rote Zeilen fuer verdaechtige Kernel
            from reportlab.platypus import TableStyle as RLTS2
            extra = []
            for i, kver in enumerate(all_kernels, 1):
                info = flags_map.get(kver, {})
                if info.get('has_antiforensics'):
                    extra.append(('BACKGROUND', (0, i), (-1, i),
                                   _rl_color_from_tuple(C_RED_L)))
                    extra.append(('TEXTCOLOR',  (0, i), (-1, i),
                                   _rl_color_from_tuple(C_RED)))
            if extra:
                tbl.setStyle(RLTS2(tbl._style._cmds + extra))
            story.append(tbl)
        elif all_kernels:
            story.append(_body(f'Kernel gefunden: {", ".join(all_kernels)}  — '
                               f'keine /boot/config-* Dateien im Image vorhanden.'))
        else:
            story.append(_body('Keine installierten Kernel erkannt.'))
        story.append(_spacer(5))

        # ── Swap-Konfiguration ────────────────────────────────────────────────────
        story.append(_h2('Swap-Konfiguration (/etc/fstab)'))
        story.append(_spacer(2))
        swap = getattr(ctx, 'swap_config', {})
        if swap.get('found'):
            entries = swap.get('entries', [])
            swap_data = [['Typ', 'Pfad / UUID', 'Groesse']]
            for e in entries:
                swap_data.append([e.get('type', '—'), e.get('path', '—'),
                                  f"{e.get('size_mb', 0):.0f} MB" if e.get('size_mb') else '—'])
            story.append(_table(swap_data, [35*mm, 110*mm, 25*mm]))
        else:
            story.append(_body(
                '❌  Kein Swap konfiguriert.  Auf Desktop-Systemen verdaechtig — '
                'RAM-Forensik aus Swap-Partition nicht moeglich.'
            ))
        story.append(_spacer(5))

        # ── Ausstehender Reboot & Kernel-Diskrepanz ───────────────────────────────
        reboot_pending_alt = getattr(ctx, 'reboot_pending', False)
        loaded_kernel_alt  = getattr(ctx, 'loaded_kernel_from_logs', '')
        active_kernel_alt  = grub.get('active_kernel', '')
        if reboot_pending_alt or (active_kernel_alt and loaded_kernel_alt
                                  and active_kernel_alt != loaded_kernel_alt):
            story.append(_h2('Kernel-Status & Reboot'))
            story.append(_spacer(2))
            state_rows = []
            state_rows.append(['GRUB-Default-Kernel', active_kernel_alt or '—'])
            state_rows.append(['Geladen laut Logs',   loaded_kernel_alt or '—'])
            if active_kernel_alt and loaded_kernel_alt and active_kernel_alt != loaded_kernel_alt:
                state_rows.append(['⚠ Kernel-Diskrepanz',
                                   f'GRUB={active_kernel_alt}  ≠  Geladen={loaded_kernel_alt}'])
            if reboot_pending_alt:
                state_rows.append(['⚠ Pending Reboot',
                                   'Kernel-Update installiert — kein Neustart erfolgt'])
            from reportlab.platypus import Table as RLT3, TableStyle as RLTS3
            data3 = [[r[0], r[1]] for r in state_rows]
            t3 = RLT3(data3, colWidths=[60*mm, W - 60*mm])
            cmds3 = [
                ('BACKGROUND',   (0, 0), (0, -1), _rl_color_from_tuple(C_LIGHT_GREY)),
                ('FONTNAME',     (0, 0), (0, -1), 'Helvetica-Bold'),
                ('FONTNAME',     (1, 0), (1, -1), 'Helvetica'),
                ('FONTSIZE',     (0, 0), (-1, -1), 8),
                ('GRID',         (0, 0), (-1, -1), 0.3, _rl_color_from_tuple((0xDD/255,)*3)),
                ('LEFTPADDING',  (0, 0), (-1, -1), 5),
                ('TOPPADDING',   (0, 0), (-1, -1), 4),
                ('BOTTOMPADDING',(0, 0), (-1, -1), 4),
            ]
            for idx, r in enumerate(state_rows):
                if '⚠' in r[0]:
                    cmds3.append(('BACKGROUND', (0, idx), (-1, idx),
                                   _rl_color_from_tuple(C_ORANGE_L)))
                    cmds3.append(('TEXTCOLOR',  (0, idx), (-1, idx),
                                   _rl_color_from_tuple(C_ORANGE)))
            t3.setStyle(RLTS3(cmds3))
            story.append(t3)
            story.append(_spacer(5))

        # ── /dev/null Symlinks ────────────────────────────────────────────────────
        symlinks_alt = getattr(ctx, 'primary_symlinks', {})
        devnull_symlinks = {p: t for p, t in symlinks_alt.items() if '/dev/null' in t}
        story.append(_h2('/dev/null Symlinks auf Log-Dateien'))
        story.append(_spacer(2))
        if devnull_symlinks:
            rows = [['Pfad', 'Symlink-Ziel', 'Bewertung']]
            for path, target in sorted(devnull_symlinks.items()):
                rows.append([path, target, '⚠ KRITISCH — keine Protokollierung'])
            tbl_sym = _table(rows, [65*mm, 55*mm, 50*mm])
            story.append(tbl_sym)
        else:
            story.append(_body('Keine /dev/null Symlinks auf Log-Dateien erkannt.'))
        story.append(_spacer(3))
        story.append(_body(
            'Methode: fls -r (Typ l/l = Symlink) + icat zum Lesen des Symlink-Ziels. '
            'Forensisch unveraendert — kein Mounting des Images.'
        ))
        story.append(PageBreak())

    # ── AUSKOMMENTIERT: Alt Seite 7 — Timeline (jetzt Seite 3) ──────────────
    # Timeline wurde auf Seite 3 vorgezogen (direkt nach Executive Summary).
    # Reaktivieren: if False entfernen.
    if False:
        story.append(_h1('Forensische Timeline'))
        story.append(_spacer(3))
        story.append(_body('Alle Zeitstempel sind in UTC.'))
        story.append(PageBreak())

    # ── AUSKOMMENTIERT: Alt Seite 8 — System-Profiling + Pipeline-Status ─────
    # System-Profiling jetzt auf Seite 7, Pipeline-Status in Anhang A.
    # Reaktivieren: if False entfernen.
    if False:
        story.append(_h1('System-Profiling & Pipeline-Status'))
        story.append(_spacer(4))
        story.append(PageBreak())

    # ── Anhang B: Parser-Statistik ────────────────────────────────────────────
    story.append(_h1('B — Parser-Statistik'))
    story.append(_spacer(3))
    story.append(_body(
        'Uebersicht aller 38 Parser + Hayabusa — '
        'welche aktiv waren und wie viele Events sie extrahiert haben.'
    ))
    story.append(_spacer(4))

    # ── AUSKOMMENTIERT: Alt Seite 9 — vollstaendiges CoC-Protokoll ───────────
    # Ersetzt durch Anhang C (Kurzform + Verweis auf chain_of_custody.pdf).
    # Reaktivieren: if False entfernen.
    if False:
        story.append(_h1('Pipeline-Ausfuehrungsprotokoll'))
        coc_tmp = ctx.coc
        if coc_tmp:
            rows = [['Stufe', 'Aktion', 'Zeitstempel']]
            for entry in coc_tmp.entries:
                rows.append([entry.stage, entry.action[:70],
                             entry.timestamp.strftime('%H:%M:%S')])
            story.append(_table(rows, [35*mm, 100*mm, 35*mm]))
        story.append(PageBreak())

    # ── Anhang B: Parser-Statistik (Inhalt) ──────────────────────────────────
    story.append(_h1('Parser-Statistik'))
    story.append(_spacer(3))
    story.append(_body('Übersicht aller 38 Parser + Hayabusa (Stage 3.3) — '
                       'welche aktiv waren und wie viele Events sie extrahiert haben.'))
    story.append(_spacer(4))

    # ── Hayabusa Info-Box ─────────────────────────────────────────────────────
    hayabusa_status = ctx.stage_status.get('stage_03_3', 'UNBEKANNT — Stage 3.3 nicht ausgeführt')
    hayabusa_hits   = ctx.hayabusa_hits

    if hayabusa_hits > 0:
        haya_bg    = C_GREEN_L
        haya_border = C_GREEN
        haya_label  = f'Hayabusa (Stage 3.3): AKTIV — {hayabusa_hits} Sigma-Treffer'
    else:
        haya_bg    = C_YELLOW_L
        haya_border = C_ORANGE
        haya_label  = f'Hayabusa (Stage 3.3): INFORMATION — {hayabusa_status}'

    haya_box = Table(
        [[Paragraph(
            f'<font size="9"><b>{haya_label}</b></font>',
            _normal_style()
        )]],
        colWidths=[W]
    )
    haya_box.setStyle(TableStyle([
        ('BACKGROUND',    (0, 0), (-1, -1), _rl_color_from_tuple(haya_bg)),
        ('BOX',           (0, 0), (-1, -1), 1.5, _rl_color_from_tuple(haya_border)),
        ('TOPPADDING',    (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING',   (0, 0), (-1, -1), 10),
    ]))
    story.append(haya_box)
    story.append(_spacer(5))

    # ── Parser-Tabelle ────────────────────────────────────────────────────────
    parser_counts = ctx.parser_stats

    all_parsers = [
        'syslog','auth','journald','kern','boot','daemon','wtmp','lastlog',
        'dpkg','apt','yum','dnf','pacman',
        'apache_access','apache_error','nginx_access','nginx_error',
        'mysql','postgresql','mongodb',
        'audit','fail2ban','ufw','cron',
        'bash_history','zsh_history','fish_history','utmp','wtmpdb',
        'postfix','ftp','samba','openvpn',
        'docker','containerd','iis','text_fallback',
        'hayabusa',
    ]
    rows = [['Parser', 'Log-Datei / Quelle', 'Events', 'Status']]
    for name in all_parsers:
        count = parser_counts.get(name, 0)
        if name == 'hayabusa':
            source = 'EVTX (Sigma-Rules)'
            status = f'Aktiv — {count} Sigma-Treffer' if count > 0 else 'Übersprungen — keine EVTX-Dateien'
        elif name == 'text_fallback' and count > 0:
            source = f'{name}.log'
            status = 'Fallback aktiv'
        else:
            source = f'{name}.log'
            status = 'Aktiv' if count > 0 else 'Nicht vorhanden'
        rows.append([name, source, str(count), status])

    story.append(_table(rows, [38*mm, 52*mm, 20*mm, 60*mm]))
    story.append(PageBreak())

    # ── Anhang C: Analyse-Protokoll (Chain of Custody) ───────────────────────
    story.append(_h1('C — Analyse-Protokoll (Chain of Custody)'))
    story.append(_spacer(3))
    story.append(_body(
        'Vollstaendiges Ausfuehrungs- und Beweismittelprotokoll mit allen Einzelschritten: '
        'chain_of_custody.pdf'
    ))
    story.append(_spacer(4))
    coc = ctx.coc
    if coc:
        kv_coc = [
            ['Dateiname', coc.file_name],
            ['SHA256',    coc.sha256],
            ['MD5',       coc.md5],
            ['Groesse',   f'{coc.size_gb:.2f} GB'],
            ['Startzeit', coc.start_time.strftime('%Y-%m-%d %H:%M:%S UTC')],
            ['Case-ID',   case_dir.name],
            ['Protokoll', f'{len(coc.entries)} Eintraege — siehe chain_of_custody.pdf'],
        ]
        story.append(_kv_table(kv_coc, W, _rl_color))
        story.append(_spacer(4))
        story.append(_body(
            'Rechtlicher Hinweis: Die Integritaet des Beweismittels wurde durch '
            'SHA256- und MD5-Hashwerte gesichert. Das vollstaendige Ausfuehrungsprotokoll '
            'ist Bestandteil der chain_of_custody.pdf.'
        ))

    # ── Anhang D: Exportierte Ausgabedateien ──────────────────────────────────
    story.append(_spacer(6))
    story.append(_h1('D — Exportierte Ausgabedateien'))
    story.append(_spacer(3))
    story.append(_body(
        'Die Pipeline generiert folgende Dateien im Case-Verzeichnis:'
    ))
    story.append(_spacer(3))
    export_files = [
        ['Datei',                      'Beschreibung'],
        ['report.pdf',                 'Dieser forensische Analysebericht (PDF)'],
        ['chain_of_custody.pdf',       'Vollstaendiges Ausfuehrungsprotokoll mit Hashwerten und Einzelschritten'],
        ['pipeline_report.json',       'Maschinenlesbare Zusammenfassung aller Stage-Ergebnisse (JSON)'],
        ['iocs.json',                  'Vollstaendige IOC-Liste (IP, Domain, Hash, E-Mail, CVE)'],
        ['antiforensics.json',         'Alle Anti-Forensik-Treffer (Typ, Datei, Details, Schwere)'],
        ['activity_timeline.csv',      'Kombinierte Event-Timeline (Logins, Reboots, Crashes) — UTC-Zeitstempel'],
        ['login_events.csv',           'Alle Login-Events (SSH, PAM, Konsole) — Benutzer, IP, Methode'],
        ['system_reboots.csv',         'Alle Reboot- und Shutdown-Ereignisse — Zeitstempel, Quelle'],
        ['system_crashes.csv',         'Kernel-Panics, OOM-Kills, Segfaults — Schwere, Quelle'],
        ['filesystem_timeline.csv',    'MACtime-Dateisystem-Timeline — MACB-Typ, Pfad, Zeitstempel'],
        ['timesketch_link.txt',        'URL zur interaktiven Timesketch-Timeline (wenn Upload erfolgreich)'],
    ]
    story.append(_table(export_files, [55*mm, 115*mm]))
    story.append(_spacer(3))

    # ── AUSKOMMENTIERT: Alt Seite 11 — YARA + Erweiterte IOCs ────────────────
    # YARA-Treffer sind Teil von ctx.antiforensics_hits (type='yara_match').
    # Erweiterte IOC-Tabelle ist Duplikat von Seite 5. Beide entfernt.
    # Reaktivieren: if False entfernen.
    if False:
        story.append(_h1('YARA-Treffer & Erweiterte IOCs'))
        yara_hits_list = [h for h in ctx.antiforensics_hits if h.get('type') == 'yara_match']
        if yara_hits_list:
            rows = [['YARA-Regel', 'Betroffene Datei', 'Details', 'Schwere']]
            for h in yara_hits_list[:20]:
                rows.append([h.get('rule', '')[:30], h['file'][-40:],
                             h['details'][:60], h['severity']])
            story.append(_table(rows, [40*mm, 55*mm, 60*mm, 15*mm]))
        story.append(_h2('Erweiterte IOC-Tabelle'))
        if ctx.iocs:
            rows = [['Typ', 'Wert', 'Parser', 'Kontext']]
            for ioc in ctx.iocs[:30]:
                rows.append([ioc.type, ioc.value[:40], ioc.source[:20], ioc.context[:40]])
            story.append(_table(rows, [22*mm, 64*mm, 25*mm, 59*mm]))

    # ── PDF bauen ─────────────────────────────────────────────────────────────
    def _on_page(canv, doc):
        _draw_header(canv, A4, case_id, created, quality)
        _draw_footer(canv, A4)

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    log.info(f'  report.pdf → {out_file}')


def _draw_header(canv, page_size, case_id, created, quality):
    from reportlab.lib.units import mm
    W, H = page_size
    canv.saveState()
    canv.setFillColorRGB(*C_DARK_BLUE)
    canv.rect(0, H - 28*mm, W, 28*mm, fill=1, stroke=0)
    canv.setFillColorRGB(*C_WHITE)
    canv.setFont('Helvetica-Bold', 13)
    canv.drawString(15*mm, H - 13*mm, 'DFIR ANALYSE-REPORT')
    canv.setFont('Helvetica', 8)
    canv.setFillColorRGB(*C_LIGHT_BLUE)
    canv.drawString(15*mm, H - 21*mm, f'Case-ID: {case_id}  |  Erstellt: {created}  |  VERTRAULICH')
    badge_color = C_GREEN if quality == 'SEHR GUT' else C_ORANGE if quality == 'GUT' else C_RED
    canv.setFillColorRGB(*badge_color)
    canv.roundRect(W - 75*mm, H - 22*mm, 60*mm, 10*mm, 2*mm, fill=1, stroke=0)
    canv.setFillColorRGB(*C_WHITE)
    canv.setFont('Helvetica-Bold', 8)
    canv.drawCentredString(W - 45*mm, H - 16*mm, f'QUALITÄT: {quality}')
    canv.restoreState()


def _draw_footer(canv, page_size):
    from reportlab.lib.units import mm
    W, H = page_size
    canv.saveState()
    canv.setStrokeColorRGB(*C_LIGHT_GREY)
    canv.setLineWidth(0.5)
    canv.line(15*mm, 18*mm, W - 15*mm, 18*mm)
    canv.setFillColorRGB(*C_MID_GREY)
    canv.setFont('Helvetica', 7)
    canv.drawString(15*mm, 12*mm,
                    'DFIR Analyse-Pipeline v3.0  |  Automatisch generiert  |  Nicht für die Öffentlichkeit')
    canv.drawRightString(W - 15*mm, 12*mm, f'Seite {canv.getPageNumber()}')
    canv.restoreState()


def _banner(text, bg, fg, size, width):
    from reportlab.platypus import Table, TableStyle
    t = Table([[text]], colWidths=[width])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), _rl_color_from_tuple(bg)),
        ('TEXTCOLOR',  (0,0), (-1,-1), _rl_color_from_tuple(fg)),
        ('FONTNAME',   (0,0), (-1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), size),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING',(0,0),(-1,-1), 8),
    ]))
    return t


def _kv_table(rows, width, rl_fn):
    """Key-Value-Tabelle mit automatischem Zeilenumbruch in der Wert-Spalte."""
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.styles import ParagraphStyle
    KEY_W = 62 * _mm_val()
    VAL_W = width - KEY_W
    _ks = ParagraphStyle('kv_k', fontSize=8, fontName='Helvetica-Bold',
                          textColor=_rl_color_from_tuple(C_DARK_GREY), leading=11)
    _vs = ParagraphStyle('kv_v', fontSize=8, fontName='Helvetica',
                          textColor=_rl_color_from_tuple(C_DARK_GREY), leading=11)
    data = [
        [Paragraph(str(r[0]), _ks), Paragraph(str(r[1]), _vs)]
        for r in rows
    ]
    t = Table(data, colWidths=[KEY_W, VAL_W])
    t.setStyle(TableStyle([
        ('BACKGROUND',   (0,0), (0,-1), _rl_color_from_tuple(C_LIGHT_GREY)),
        ('BACKGROUND',   (1,0), (1,-1), _rl_color_from_tuple(C_WHITE)),
        ('GRID',         (0,0), (-1,-1), 0.3, _rl_color_from_tuple((0xDD/255,)*3)),
        ('LEFTPADDING',  (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING',   (0,0), (-1,-1), 5),
        ('BOTTOMPADDING',(0,0), (-1,-1), 5),
        ('VALIGN',       (0,0), (-1,-1), 'TOP'),
    ]))
    return t


def _stat_boxes(stats, width, rl_fn):
    from reportlab.platypus import Table, TableStyle, Paragraph
    n    = len(stats)
    cell_w = width / n
    data = [[Paragraph(f'<font size="22" color="#{_hex(C_MID_BLUE)}"><b>{v}</b></font>'
                       f'<br/><font size="8" color="#{_hex(C_MID_GREY)}">{label}</font>',
                       _normal_style())
             for v, label in stats]]
    t = Table(data, colWidths=[cell_w]*n)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,-1), _rl_color_from_tuple(C_LIGHT_GREY)),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING',(0,0),(-1,-1), 8),
    ]))
    return t


def _rl_color_from_tuple(t):
    from reportlab.lib.colors import Color
    return Color(*t)


def _hex(rgb_tuple) -> str:
    return ''.join(f'{int(v*255):02X}' for v in rgb_tuple)


def _mm_val():
    from reportlab.lib.units import mm
    return mm


def _normal_style():
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.enums import TA_CENTER
    styles = getSampleStyleSheet()
    s = styles['Normal'].clone('centered')
    s.alignment = TA_CENTER
    return s


def _normal_style_left():
    """Links-ausgerichteter Paragraph-Style — fuer Befund-Boxen."""
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.enums import TA_LEFT
    styles = getSampleStyleSheet()
    s = styles['Normal'].clone('left_aligned')
    s.alignment = TA_LEFT
    return s


# ── Chain of Custody PDF ──────────────────────────────────────────────────────

def _generate_coc_pdf(ctx: PipelineContext, case_dir: Path) -> None:
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer,
                                        Table, TableStyle, PageBreak)
        from reportlab.lib.styles import getSampleStyleSheet
    except ImportError:
        return

    out_file = case_dir / 'chain_of_custody.pdf'
    doc      = SimpleDocTemplate(str(out_file), pagesize=A4,
                                  leftMargin=15*mm, rightMargin=15*mm,
                                  topMargin=35*mm, bottomMargin=22*mm)
    styles   = getSampleStyleSheet()
    story    = []
    coc      = ctx.coc
    case_id  = case_dir.name
    created  = datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')
    quality  = evaluate_quality(ctx)

    W = A4[0] - 30*mm

    story.append(_banner('CHAIN OF CUSTODY', C_DARK_BLUE, C_WHITE, 20, W))
    story.append(Spacer(1, 6*mm))

    if coc:
        kv = [
            ['Dateiname', coc.file_name],
            ['SHA256',    coc.sha256],
            ['MD5',       coc.md5],
            ['Größe',     f'{coc.size_gb:.2f} GB'],
            ['Startzeit', coc.start_time.strftime('%Y-%m-%d %H:%M:%S UTC')],
            ['Case-ID',   case_id],
        ]
        data = [[r[0], r[1]] for r in kv]
        t = Table(data, colWidths=[50*mm, W - 50*mm])
        t.setStyle(TableStyle([
            ('BACKGROUND',  (0,0), (0,-1), _rl_color_from_tuple(C_LIGHT_GREY)),
            ('FONTNAME',    (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTNAME',    (1,0), (1,-1), 'Helvetica'),
            ('FONTSIZE',    (0,0), (-1,-1), 9),
            ('GRID',        (0,0), (-1,-1), 0.3, _rl_color_from_tuple((0xDD/255,)*3)),
            ('LEFTPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING',  (0,0), (-1,-1), 4),
            ('BOTTOMPADDING',(0,0),(-1,-1), 4),
        ]))
        story.append(t)
        story.append(Spacer(1, 6*mm))

        rows = [['Stufe', 'Aktion', 'Zeitstempel']]
        for entry in coc.entries:
            rows.append([entry.stage, entry.action[:80],
                         entry.timestamp.strftime('%Y-%m-%d %H:%M:%S')])
        t2 = Table(rows, colWidths=[35*mm, 105*mm, 30*mm], repeatRows=1)
        t2.setStyle(TableStyle([
            ('BACKGROUND',  (0,0), (-1,0), _rl_color_from_tuple(C_DARK_BLUE)),
            ('TEXTCOLOR',   (0,0), (-1,0), _rl_color_from_tuple(C_WHITE)),
            ('FONTNAME',    (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTNAME',    (0,1), (-1,-1), 'Helvetica'),
            ('FONTSIZE',    (0,0), (-1,-1), 8),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),
             [_rl_color_from_tuple(C_WHITE), _rl_color_from_tuple(C_LIGHT_GREY)]),
            ('GRID',        (0,0), (-1,-1), 0.3, _rl_color_from_tuple((0xDD/255,)*3)),
            ('LEFTPADDING', (0,0), (-1,-1), 5),
            ('TOPPADDING',  (0,0), (-1,-1), 4),
            ('BOTTOMPADDING',(0,0),(-1,-1), 4),
        ]))
        story.append(t2)

        # Extrahierte Dateien + Hashwerte
        if coc.extracted_file_hashes:
            story.append(Spacer(1, 6*mm))
            story.append(_banner(
                f'EXTRAHIERTE DATEIEN — HASHWERTE ({len(coc.extracted_file_hashes)} Dateien)',
                C_MID_BLUE, C_WHITE, 11, W))
            story.append(Spacer(1, 3*mm))
            hash_rows = [['Datei (Pfad)', 'SHA256', 'MD5']]
            for fname, h in sorted(coc.extracted_file_hashes.items()):
                if isinstance(h, dict):
                    sha, md5 = h.get('sha256', ''), h.get('md5', '')
                else:                       # Alt-Format (Snapshot-Reexport)
                    sha, md5 = h, ''
                # Pfad ggf. links kuerzen — Hashes NIEMALS kuerzen
                hash_rows.append([('…' + fname[-48:]) if len(fname) > 49 else fname,
                                  sha, md5])
            t3 = Table(hash_rows,
                       colWidths=[52*mm, (W - 52*mm) * 0.62, (W - 52*mm) * 0.38],
                       repeatRows=1)
            t3.setStyle(TableStyle([
                ('BACKGROUND',    (0,0), (-1,0), _rl_color_from_tuple(C_MID_BLUE)),
                ('TEXTCOLOR',     (0,0), (-1,0), _rl_color_from_tuple(C_WHITE)),
                ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
                ('FONTNAME',      (0,1), (-1,-1), 'Helvetica'),
                ('FONTSIZE',      (0,0), (-1,-1), 7),
                ('ROWBACKGROUNDS',(0,1), (-1,-1),
                 [_rl_color_from_tuple(C_WHITE), _rl_color_from_tuple(C_LIGHT_GREY)]),
                ('GRID',          (0,0), (-1,-1), 0.3, _rl_color_from_tuple((0xDD/255,)*3)),
                ('LEFTPADDING',   (0,0), (-1,-1), 5),
                ('TOPPADDING',    (0,0), (-1,-1), 3),
                ('BOTTOMPADDING', (0,0), (-1,-1), 3),
            ]))
            story.append(t3)

    def _on_page(canv, doc):
        _draw_header(canv, A4, case_id, created, quality)
        _draw_footer(canv, A4)

    doc.build(story, onFirstPage=_on_page, onLaterPages=_on_page)
    log.info(f'  chain_of_custody.pdf → {out_file}')


# ── Timesketch Upload ─────────────────────────────────────────────────────────

def _upload_timesketch(ctx: PipelineContext, case_dir: Path) -> None:
    try:
        import yaml
        cfg_path = Path(__file__).parent.parent / 'config.yaml'
        if not cfg_path.exists():
            _write_timesketch_link(case_dir, 'http://localhost:5000')
            return
        with open(cfg_path) as f:
            cfg = yaml.safe_load(f)
        ts_cfg  = cfg.get('timesketch', {})
        host    = ts_cfg.get('host', 'http://localhost:5000')
        user    = ts_cfg.get('username', 'admin')
        passwd  = ts_cfg.get('password', 'changeme')
        sketch  = ts_cfg.get('sketch_id', 1)

        from timesketch_api_client import client as ts_client
        cli    = ts_client.TimesketchApi(host, user, passwd)
        sk     = cli.get_sketch(sketch)
        events = []
        for e in ctx.normalized_events[:50000]:
            events.append({
                'datetime':    e.timestamp.isoformat(),
                'timestamp_desc': e.event_type,
                'message':     e.message,
                'source':      e.source,
            })
        sk.add_timeline_from_json(json.dumps(events),
                                  timeline_name=f'DFIR_{case_dir.name}')
        url = f'{host}/sketch/{sketch}/explore'
        log.info(f'  Timesketch Upload abgeschlossen → {url}')
        _write_timesketch_link(case_dir, url)

    except Exception as e:
        log.warning(f'  Timesketch Upload fehlgeschlagen: {e}')
        _write_timesketch_link(case_dir, 'http://localhost:5000')


def _write_timesketch_link(case_dir: Path, url: str) -> None:
    (case_dir / 'timesketch_link.txt').write_text(
        f'Timesketch URL: {url}\n'
        f'Erstellt: {datetime.utcnow().isoformat()}Z\n'
    )


try:
    from reportlab.platypus import Paragraph as _Paragraph
except ImportError:
    pass
